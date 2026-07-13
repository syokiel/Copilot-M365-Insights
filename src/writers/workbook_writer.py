from pathlib import Path

import openpyxl

from src.writers import (
    sheet_agents,
    sheet_ai_usage,
    sheet_az_health,
    sheet_connectors,
    sheet_crossref,
    sheet_dlp,
    sheet_environments,
    sheet_invocations,
    sheet_kpi_history,
    sheet_m365_admin_inventory,
    sheet_m365_copilot,
    sheet_m365_copilot_trend,
    sheet_m365_packages,
    sheet_m365_o365_users,
    sheet_m365_app_users,
    sheet_m365_usage_agents,
    sheet_m365_usage_agent_users,
    sheet_m365_usage_users,
    sheet_publishers,
    sheet_summary,
    sheet_teams_usage,
    sheet_tokenomics_capacity,
    sheet_tokenomics_entitlement,
    sheet_tokenomics_entitlement_per_agent,
    sheet_tokenomics_entitlement_per_user,
    sheet_tokenomics_summary,
    sheet_viva,
    sheet_viva_adoption,
    sheet_viva_impact,
    sheet_viva_sessions,
    sheet_viva_topics,
    sheet_viva_wau,
    sheet_viva_autonomous,
    sheet_xla,
    sheet_xla_persona_journey,
    sheet_xla_agent_contribution,
    sheet_m365_licence_optimization,
    sheet_m365_f3_candidates,
    sheet_m365_activations,
    sheet_m365_services_counts,
    sheet_m365_activity_counts,
    sheet_m365_active_user_counts,
    sheet_m365_active_users_detail,
    sheet_m365_proplus_platforms,
    sheet_m365_proplus_counts,
    sheet_m365_proplus_detail,
    sheet_billing_licences,
)


def build_workbook(
    events: list[dict],
    connector_calls: list[dict],
    output_path: str,
    agents: list[dict] | None = None,
    environments: list[dict] | None = None,
    publishers: list[dict] | None = None,
    dlp_policies: list[dict] | None = None,
    agent_solutions: list[dict] | None = None,
    aad_users: dict[str, dict] | None = None,
    model_calls: list[dict] | None = None,
    health_detail: list[dict] | None = None,
    crossref_summary: list[dict] | None = None,
    copilot_usage: list[dict] | None = None,
    teams_usage: list[dict] | None = None,
    kpi_snapshots: list[dict] | None = None,
    viva_person_insights: list[dict] | None = None,
    viva_reports_cs_session_metrics: list[dict] | None = None,
    viva_reports_cs_topic_metrics: list[dict] | None = None,
    viva_reports_cs_weekly_active_users: list[dict] | None = None,
    viva_reports_cs_autonomous_metrics: list[dict] | None = None,
    viva_reports_cs_copilot_agents: dict[str, dict] | None = None,
    copilot_count_summary: list[dict] | None = None,
    copilot_count_trend: list[dict] | None = None,
    copilot_packages: list[dict] | None = None,
    o365_active_users: list[dict] | None = None,
    m365_app_users: list[dict] | None = None,
    viva_reports_copilot_adoption: list[dict] | None = None,
    viva_reports_copilot_impact: list[dict] | None = None,
    m365_admin_agent_inventory: list[dict] | None = None,
    m365_usage_agents: list[dict] | None = None,
    m365_usage_agent_users: list[dict] | None = None,
    m365_usage_users: list[dict] | None = None,
    viva_reports_cs_action_metrics: list[dict] | None = None,
    tokenomics_capacity_consumption: list[dict] | None = None,
    tokenomics_entitlement_consumption: list[dict] | None = None,
    tokenomics_entitlement_per_agent: list[dict] | None = None,
    tokenomics_entitlement_per_user: list[dict] | None = None,
    xla_by_persona_journey: list[dict] | None = None,
    xla_agent_contribution: list[dict] | None = None,
    m365_usage_activations_users: list[dict] | None = None,
    m365_usage_active_users_services: list[dict] | None = None,
    m365_usage_active_users_activity: list[dict] | None = None,
    m365_usage_active_user_counts: list[dict] | None = None,
    m365_usage_active_users_detail: list[dict] | None = None,
    m365_usage_proplus_platforms: list[dict] | None = None,
    m365_usage_proplus_counts: list[dict] | None = None,
    m365_usage_proplus_detail: list[dict] | None = None,
    billing_licences: list[dict] | None = None,
) -> None:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    def _if(name: str, write_fn, *args):
        """Create the sheet only when at least one data arg is non-empty."""
        has_data = any(
            (isinstance(a, (list, tuple)) and len(a) > 0) or
            (isinstance(a, dict) and len(a) > 0)
            for a in args
        )
        if has_data:
            write_fn(wb.create_sheet(name), *args)

    latest_kpi = kpi_snapshots[0] if kpi_snapshots else None

    # Always-present sheets (summary / history depend on all data)
    sheet_kpi_history.write(
        wb.create_sheet("KPI History"),
        kpi_snapshots or [],
        entitlement=tokenomics_entitlement_consumption or [],
        per_agent=tokenomics_entitlement_per_agent or [],
        capacity=tokenomics_capacity_consumption or [],
    )
    sheet_summary.write(
        wb.create_sheet("Copilot_Adoption_Summary"), events, connector_calls, model_calls or [],
        kpi_snapshot=latest_kpi,
        viva_reports_cs_sessions=viva_reports_cs_session_metrics or [],
        viva_reports_cs_wau=viva_reports_cs_weekly_active_users or [],
        viva_reports_cs_autonomous=viva_reports_cs_autonomous_metrics or [],
        viva_reports_cs_agents=viva_reports_cs_copilot_agents or {},
    )

    # XLA Scorecard — only when there's session data to compute from
    _xla_sess = viva_reports_cs_session_metrics or []
    if _xla_sess:
        sheet_xla.write(
            wb.create_sheet("XLA_Measurements"),
            session_metrics=_xla_sess,
            autonomous_metrics=viva_reports_cs_autonomous_metrics or [],
            action_metrics=viva_reports_cs_action_metrics or [],
            wau=viva_reports_cs_weekly_active_users or [],
            adoption=viva_reports_copilot_adoption or [],
            kpi_snapshot=latest_kpi,
            agents=viva_reports_cs_copilot_agents or {},
        )

    _tok = (
        tokenomics_entitlement_consumption or [],
        tokenomics_entitlement_per_agent or [],
        tokenomics_entitlement_per_user or [],
        tokenomics_capacity_consumption or [],
    )
    if any(_tok):
        sheet_tokenomics_summary.write(
            wb.create_sheet("Tokenomics_Summary"),
            entitlement=tokenomics_entitlement_consumption or [],
            per_agent=tokenomics_entitlement_per_agent or [],
            per_user=tokenomics_entitlement_per_user or [],
            capacity=tokenomics_capacity_consumption or [],
        )

    if billing_licences:
        sheet_m365_licence_optimization.write(
            wb.create_sheet("M365_Licence_Optimization"),
            billing_licences=billing_licences or [],
            services_counts=m365_usage_active_users_services or [],
            active_users_detail=m365_usage_active_users_detail or [],
            activations_users=m365_usage_activations_users or [],
            proplus_counts=m365_usage_proplus_counts or [],
        )

    if m365_usage_active_users_detail and m365_usage_proplus_detail:
        sheet_m365_f3_candidates.write(
            wb.create_sheet("M365_F3_Candidates"),
            active_users_detail=m365_usage_active_users_detail,
            proplus_detail=m365_usage_proplus_detail,
        )

    _if("Invocations",           sheet_invocations.write,         events, connector_calls)
    _if("Connectors",            sheet_connectors.write,           connector_calls)
    _if("AI_Model_Calls",        sheet_ai_usage.write,             model_calls or [])
    _if("Agents",                sheet_agents.write,               agents or [], environments or [], agent_solutions or [], aad_users or {})
    _if("Environments",          sheet_environments.write,         environments or [])
    _if("Publishers",            sheet_publishers.write,           publishers or [])
    _if("DLP Policies",          sheet_dlp.write,                  dlp_policies or [])
    _if("M365_Copilot_Usage",    sheet_m365_copilot.write,         copilot_usage or [])
    _if("M365_Copilot_Trend",    sheet_m365_copilot_trend.write,   copilot_count_summary or [], copilot_count_trend or [])
    _if("M365_Copilot_Packages", sheet_m365_packages.write,        copilot_packages or [])
    _if("M365_O365_Users",       sheet_m365_o365_users.write,      o365_active_users or [])
    _if("M365_App_Users",        sheet_m365_app_users.write,       m365_app_users or [])
    _if("Teams_Usage",           sheet_teams_usage.write,          teams_usage or [])
    _if("Viva_Person_Insights",  sheet_viva.write,                 viva_person_insights or [], aad_users or {})
    _if("Viva_CS_Sessions",      sheet_viva_sessions.write,        viva_reports_cs_session_metrics or [], viva_reports_cs_copilot_agents or {})
    _if("Viva_CS_Topics",        sheet_viva_topics.write,          viva_reports_cs_topic_metrics or [], viva_reports_cs_copilot_agents or {})
    _if("Viva_CS_WAU",           sheet_viva_wau.write,             viva_reports_cs_weekly_active_users or [], viva_reports_cs_copilot_agents or {})
    _if("Viva_CS_Autonomous",    sheet_viva_autonomous.write,      viva_reports_cs_autonomous_metrics or [], viva_reports_cs_copilot_agents or {})
    _if("Viva_Copilot_Adoption", sheet_viva_adoption.write,        viva_reports_copilot_adoption or [])
    _if("Viva_Copilot_Impact",   sheet_viva_impact.write,          viva_reports_copilot_impact or [])
    _if("M365_Agent_Inventory",  sheet_m365_admin_inventory.write, m365_admin_agent_inventory or [])
    _if("M365_Usage_Agents",     sheet_m365_usage_agents.write,    m365_usage_agents or [])
    _if("M365_Usage_AgentUsers", sheet_m365_usage_agent_users.write, m365_usage_agent_users or [])
    _if("M365_Usage_Users",      sheet_m365_usage_users.write,     m365_usage_users or [])
    _if("Tokenomics_Capacity",    sheet_tokenomics_capacity.write,    tokenomics_capacity_consumption or [])
    _if("Tokenomics_Entitlement", sheet_tokenomics_entitlement.write, tokenomics_entitlement_consumption or [])
    _if("Tokenomics_PerAgent",    sheet_tokenomics_entitlement_per_agent.write, tokenomics_entitlement_per_agent or [])
    _if("Tokenomics_PerUser",     sheet_tokenomics_entitlement_per_user.write,  tokenomics_entitlement_per_user or [])
    _if("XLA_Persona_Journey",    sheet_xla_persona_journey.write,    xla_by_persona_journey or [])
    _if("XLA_Agent_Contribution", sheet_xla_agent_contribution.write, xla_agent_contribution or [])
    _if("M365_Activations",        sheet_m365_activations.write,         m365_usage_activations_users or [])
    _if("M365_Services_Counts",   sheet_m365_services_counts.write,     m365_usage_active_users_services or [])
    _if("M365_Activity_Counts",   sheet_m365_activity_counts.write,     m365_usage_active_users_activity or [])
    _if("M365_Active_Counts",     sheet_m365_active_user_counts.write,  m365_usage_active_user_counts or [])
    _if("M365_Active_Users",      sheet_m365_active_users_detail.write, m365_usage_active_users_detail or [])
    _if("M365_ProPlus_Platforms", sheet_m365_proplus_platforms.write,   m365_usage_proplus_platforms or [])
    _if("M365_ProPlus_Counts",    sheet_m365_proplus_counts.write,      m365_usage_proplus_counts or [])
    _if("M365_ProPlus_Users",     sheet_m365_proplus_detail.write,      m365_usage_proplus_detail or [])
    _if("Billing_Licences",       sheet_billing_licences.write,         billing_licences or [])
    _if("AzureMonitor_Health",    sheet_az_health.write,                health_detail or [])
    _if("CrossRef_Summary",       sheet_crossref.write,                 crossref_summary or [])

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"Saved: {output_path}")
