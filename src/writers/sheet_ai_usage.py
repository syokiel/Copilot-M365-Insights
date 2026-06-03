from collections import defaultdict

from openpyxl.worksheet.worksheet import Worksheet

from src.writers._style import apply_row_style, autofit_columns, write_headers

HEADERS = [
    "Timestamp",
    "gen_ai.operation.name",
    "gen_ai.provider.name",
    "gen_ai.request.model",
    "gen_ai.response.model",
    "gen_ai.usage.input_tokens",
    "gen_ai.usage.output_tokens",
    "Total Tokens",
    "gen_ai.agent.name",
    "gen_ai.agent.id",
    "gen_ai.environment.id",
    "Conversation ID",
    "User ID",
    "Duration (ms)",
    "Success",
    "Dependency Type",
    "Target",
]

SUMMARY_HEADERS = [
    "gen_ai.agent.name",
    "gen_ai.agent.id",
    "Model(s) Used",
    "Provider(s)",
    "Total Calls",
    "Total Input Tokens",
    "Total Output Tokens",
    "Total Tokens",
]


def write(ws: Worksheet, model_calls: list[dict]) -> None:
    """Two sections: per-agent summary at top, per-call detail below."""
    if not model_calls:
        write_headers(ws, HEADERS)
        ws.cell(row=2, column=1, value="— No AI model call data found in the lookback window —")
        autofit_columns(ws, HEADERS)
        return

    # ── Summary by agent ──────────────────────────────────────────────────────
    by_agent: dict[str, dict] = defaultdict(lambda: {
        "agent_name": "", "agent_id": "", "models": set(), "providers": set(),
        "calls": 0, "input_tokens": 0, "output_tokens": 0,
    })

    for r in model_calls:
        agent_id = r.get("gen_ai_agent_id") or "unknown"
        a = by_agent[agent_id]
        a["agent_name"] = a["agent_name"] or r.get("gen_ai_agent_name", "")
        a["agent_id"] = agent_id
        model = r.get("gen_ai_request_model") or r.get("gen_ai_response_model", "")
        if model:
            a["models"].add(model)
        provider = r.get("gen_ai_provider_name", "")
        if provider:
            a["providers"].add(provider)
        a["calls"] += 1
        a["input_tokens"] += r.get("gen_ai_usage_input_tokens") or 0
        a["output_tokens"] += r.get("gen_ai_usage_output_tokens") or 0

    summary_rows = sorted(by_agent.values(), key=lambda x: x["calls"], reverse=True)

    write_headers(ws, SUMMARY_HEADERS)
    for row_idx, a in enumerate(summary_rows, 2):
        ws.cell(row=row_idx, column=1, value=a["agent_name"] or a["agent_id"])
        ws.cell(row=row_idx, column=2, value=a["agent_id"])
        ws.cell(row=row_idx, column=3, value=", ".join(sorted(a["models"])))
        ws.cell(row=row_idx, column=4, value=", ".join(sorted(a["providers"])))
        ws.cell(row=row_idx, column=5, value=a["calls"])
        ws.cell(row=row_idx, column=6, value=a["input_tokens"])
        ws.cell(row=row_idx, column=7, value=a["output_tokens"])
        ws.cell(row=row_idx, column=8, value=a["input_tokens"] + a["output_tokens"])
        apply_row_style(ws, row_idx, len(SUMMARY_HEADERS))

    # ── Per-call detail ───────────────────────────────────────────────────────
    detail_start = len(summary_rows) + 4
    write_headers(ws, HEADERS, start_row=detail_start)

    calls_sorted = sorted(model_calls, key=lambda r: r.get("timestamp") or "", reverse=True)
    for row_idx, r in enumerate(calls_sorted, detail_start + 1):
        input_tok = r.get("gen_ai_usage_input_tokens") or 0
        output_tok = r.get("gen_ai_usage_output_tokens") or 0
        ws.cell(row=row_idx, column=1, value=(r.get("timestamp") or "")[:19])
        ws.cell(row=row_idx, column=2, value=r.get("gen_ai_operation_name", ""))
        ws.cell(row=row_idx, column=3, value=r.get("gen_ai_provider_name", ""))
        ws.cell(row=row_idx, column=4, value=r.get("gen_ai_request_model", ""))
        ws.cell(row=row_idx, column=5, value=r.get("gen_ai_response_model", ""))
        ws.cell(row=row_idx, column=6, value=input_tok or None)
        ws.cell(row=row_idx, column=7, value=output_tok or None)
        ws.cell(row=row_idx, column=8, value=(input_tok + output_tok) or None)
        ws.cell(row=row_idx, column=9, value=r.get("gen_ai_agent_name", ""))
        ws.cell(row=row_idx, column=10, value=r.get("gen_ai_agent_id", ""))
        ws.cell(row=row_idx, column=11, value=r.get("gen_ai_environment_id", ""))
        ws.cell(row=row_idx, column=12, value=r.get("conversation_id", ""))
        ws.cell(row=row_idx, column=13, value=r.get("user_id", ""))
        ws.cell(row=row_idx, column=14, value=r.get("duration_ms"))
        ws.cell(row=row_idx, column=15, value="Yes" if r.get("success") else "No")
        ws.cell(row=row_idx, column=16, value=r.get("dependency_type", ""))
        ws.cell(row=row_idx, column=17, value=r.get("target", ""))
        apply_row_style(ws, row_idx, len(HEADERS))

    autofit_columns(ws, HEADERS)
