import re

from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from src.writers._style import DATA_FONT, HEADER_FILL, HEADER_FONT, LEFT

_INPUT_FILL   = PatternFill("solid", fgColor="FFF2CC")  # yellow  — user-editable
_TOTAL_FILL   = PatternFill("solid", fgColor="E2EFDA")  # green   — calculated totals
_SECTION_FONT = Font(bold=True, size=11, color="1F4E79")
_BOLD         = Font(bold=True, size=11)

# ── Fixed row addresses used by Excel formulas ────────────────────────────────
_PRICE_F3_ROW    = 3
_PRICE_E3_ROW    = 4
_PRICE_E5_ROW    = 5
_PRICE_E7_ROW    = 6
_COUNT_E3_ROW    = 11
_COUNT_E5_ROW    = 12
_COUNT_E7_ROW    = 13
_COUNT_TOTAL_ROW = 14
_SAV_E3_ROW      = 16
_SAV_E5_ROW      = 17
_SAV_E7_ROW      = 18
_SAV_MONTHLY_ROW = 19
_SAV_ANNUAL_ROW  = 20
_CAVEAT_START    = 23
_LIST_HEADER_ROW = 31
_DATA_START_ROW  = 32


def _sku_tier(products: str) -> str:
    for sku in ["E7", "E5", "E3"]:
        if re.search(r"\b" + sku + r"\b", products.upper()):
            return sku
    return ""


def _sku_names(products: str) -> str:
    parts = [p.strip() for p in products.split("+")]
    e_parts = [p for p in parts if re.search(r"\b(E3|E5|E7)\b", p.upper())]
    return " + ".join(e_parts) if e_parts else ""


def write(
    ws: Worksheet,
    active_users_detail: list[dict],
    proplus_detail: list[dict],
) -> None:

    # ── Build candidate list ──────────────────────────────────────────────────
    pp = {r["user_principal_name"]: r for r in proplus_detail}
    candidates: list[dict] = []
    counts = {"E3": 0, "E5": 0, "E7": 0}

    for d in active_users_detail:
        if d.get("is_deleted"):
            continue
        tier = _sku_tier(d.get("assigned_products", ""))
        if not tier:
            continue
        p = pp.get(d["user_principal_name"])
        if not p:
            continue  # no ProPlus record — cannot confirm web/mobile-only usage
        if p.get("windows") or p.get("mac"):
            continue  # uses desktop Office — not a safe F3 candidate
        counts[tier] += 1
        candidates.append({
            "user_principal_name":  d["user_principal_name"],
            "display_name":         d.get("display_name", ""),
            "tier":                 tier,
            "license_skus":         _sku_names(d.get("assigned_products", "")),
            "mobile":               p.get("mobile", 0),
            "web":                  p.get("web", 0),
            "outlook":              p.get("outlook", 0),
            "teams":                p.get("teams", 0),
            "teams_last_activity":    d.get("teams_last_activity", ""),
            "exchange_last_activity": d.get("exchange_last_activity", ""),
            "last_activity_date":     p.get("last_activity_date", ""),
        })

    candidates.sort(key=lambda r: (r["tier"], r["user_principal_name"]))

    # ── Helper ────────────────────────────────────────────────────────────────
    def _cell(row, col, value=None, *, bold=False, fill=None, fmt=None, merge_to=None):
        if merge_to:
            ws.merge_cells(f"{get_column_letter(col)}{row}:{get_column_letter(merge_to)}{row}")
        c = ws.cell(row=row, column=col, value=value)
        c.alignment = LEFT
        c.font = _BOLD if bold else DATA_FONT
        if fill:
            c.fill = fill
        if fmt:
            c.number_format = fmt
        return c

    # ═══════════════════════════════════════════════════════════════════════════
    # SECTION 1 — SKU Price Configuration
    # ═══════════════════════════════════════════════════════════════════════════
    _cell(1, 1, "SKU Price Configuration  —  Edit the highlighted 'Monthly Cost' cells to update savings estimates",
          fill=HEADER_FILL, merge_to=3).font = HEADER_FONT

    for col, hdr in enumerate(["SKU", "Monthly Cost ($/user)", "Description"], 1):
        c = ws.cell(row=2, column=col, value=hdr)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = LEFT

    price_data = [
        (_PRICE_F3_ROW, "F3",  8.00, "Microsoft 365 F3 — Frontline Worker  (web + mobile only, 2 GB mailbox)"),
        (_PRICE_E3_ROW, "E3", 36.00, "Microsoft 365 E3 — Enterprise"),
        (_PRICE_E5_ROW, "E5", 57.00, "Microsoft 365 E5 — Enterprise + Security & Compliance"),
        (_PRICE_E7_ROW, "E7",  0.00, "Custom SKU — enter your contract price"),
    ]
    for row, sku, price, desc in price_data:
        _cell(row, 1, sku)
        pc = ws.cell(row=row, column=2, value=price)
        pc.fill = _INPUT_FILL
        pc.font = _BOLD
        pc.number_format = '$#,##0.00'
        pc.alignment = LEFT
        _cell(row, 3, desc)

    # ═══════════════════════════════════════════════════════════════════════════
    # SECTION 2 — Candidate Summary
    # ═══════════════════════════════════════════════════════════════════════════
    _cell(9, 1, "── F3 Downgrade Candidate Summary ──────────────────────────────────────",
          fill=None, merge_to=3).font = _SECTION_FONT

    for col, hdr in enumerate(["Metric", "Count / Value"], 1):
        c = ws.cell(row=10, column=col, value=hdr)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = LEFT

    for row, tier in [(_COUNT_E3_ROW, "E3"), (_COUNT_E5_ROW, "E5"), (_COUNT_E7_ROW, "E7")]:
        _cell(row, 1, f"{tier} → F3 candidates  (no desktop app usage in period)")
        _cell(row, 2, counts[tier])

    _cell(_COUNT_TOTAL_ROW, 1, "Total F3 candidates", bold=True, fill=_TOTAL_FILL)
    tc = ws.cell(row=_COUNT_TOTAL_ROW, column=2,
                 value=f"=B{_COUNT_E3_ROW}+B{_COUNT_E5_ROW}+B{_COUNT_E7_ROW}")
    tc.font = _BOLD
    tc.fill = _TOTAL_FILL
    tc.alignment = LEFT

    # Savings rows (all formula-driven — react to price cell edits)
    savings_rows = [
        (_SAV_E3_ROW, "Est. monthly savings — E3 → F3",
         f"=B{_COUNT_E3_ROW}*(B{_PRICE_E3_ROW}-B{_PRICE_F3_ROW})", False),
        (_SAV_E5_ROW, "Est. monthly savings — E5 → F3",
         f"=B{_COUNT_E5_ROW}*(B{_PRICE_E5_ROW}-B{_PRICE_F3_ROW})", False),
        (_SAV_E7_ROW, "Est. monthly savings — E7 → F3",
         f"=B{_COUNT_E7_ROW}*(B{_PRICE_E7_ROW}-B{_PRICE_F3_ROW})", False),
        (_SAV_MONTHLY_ROW, "Total estimated monthly savings",
         f"=B{_SAV_E3_ROW}+B{_SAV_E5_ROW}+B{_SAV_E7_ROW}", True),
        (_SAV_ANNUAL_ROW, "Total estimated annual savings",
         f"=B{_SAV_MONTHLY_ROW}*12", True),
    ]
    for row, label, formula, is_total in savings_rows:
        fill = _TOTAL_FILL if is_total else None
        lc = ws.cell(row=row, column=1, value=label)
        lc.font = _BOLD if is_total else DATA_FONT
        lc.alignment = LEFT
        if fill:
            lc.fill = fill
        vc = ws.cell(row=row, column=2, value=formula)
        vc.font = _BOLD if is_total else DATA_FONT
        vc.number_format = "$#,##0.00"
        vc.alignment = LEFT
        if fill:
            vc.fill = fill

    # ═══════════════════════════════════════════════════════════════════════════
    # SECTION 3 — Caveats
    # ═══════════════════════════════════════════════════════════════════════════
    _cell(_CAVEAT_START, 1,
          "── Caveats — Review before migrating ───────────────────────────────────",
          merge_to=3).font = _SECTION_FONT

    caveats = [
        ("Mailbox size",   "F3 includes a 2 GB mailbox vs 100 GB on E3/E5 — verify mailbox sizes before migrating"),
        ("Desktop apps",   "Candidates are users with zero desktop app usage in this report period only — verify historical patterns"),
        ("Feature gaps",   "F3 excludes advanced compliance, Intune, Phone System, and some Power Platform capabilities"),
        ("ProPlus data",   "Users with no ProPlus report record are excluded from candidates as usage cannot be confirmed"),
        ("Pricing",        "Default prices are Microsoft list prices — update the highlighted cells above with your contract pricing"),
    ]
    for i, (label, note) in enumerate(caveats, start=_CAVEAT_START + 1):
        ws.cell(row=i, column=1, value=label).font = _BOLD
        ws.cell(row=i, column=1).alignment = LEFT
        ws.cell(row=i, column=3, value=note).font = DATA_FONT
        ws.cell(row=i, column=3).alignment = LEFT

    # ═══════════════════════════════════════════════════════════════════════════
    # SECTION 4 — Per-user candidate list
    # ═══════════════════════════════════════════════════════════════════════════
    _cell(_LIST_HEADER_ROW - 1, 1,
          "── Per-User F3 Candidates ──────────────────────────────────────────────",
          merge_to=11).font = _SECTION_FONT

    list_headers = [
        "User Principal Name", "Display Name", "Tier", "E-SKU License(s)",
        "Mobile", "Web", "Outlook Active", "Teams Active",
        "Teams Last Activity", "Exchange Last Activity", "ProPlus Last Activity",
    ]
    for col, hdr in enumerate(list_headers, 1):
        c = ws.cell(row=_LIST_HEADER_ROW, column=col, value=hdr)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = LEFT

    ws.freeze_panes = f"A{_DATA_START_ROW}"
    last_data_row = max(_DATA_START_ROW, _DATA_START_ROW + len(candidates) - 1)
    ws.auto_filter.ref = (
        f"A{_LIST_HEADER_ROW}:{get_column_letter(len(list_headers))}{last_data_row}"
    )

    _ODD  = PatternFill("solid", fgColor="FFFFFF")
    _EVEN = PatternFill("solid", fgColor="D6E4F0")

    for i, r in enumerate(candidates, start=_DATA_START_ROW):
        fill = _EVEN if i % 2 == 0 else _ODD
        vals = [
            r["user_principal_name"], r["display_name"], r["tier"], r["license_skus"],
            "Yes" if r["mobile"] else "No",
            "Yes" if r["web"]    else "No",
            "Yes" if r["outlook"] else "No",
            "Yes" if r["teams"]  else "No",
            r["teams_last_activity"], r["exchange_last_activity"], r["last_activity_date"],
        ]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=i, column=col, value=val)
            c.fill = fill
            c.font = DATA_FONT
            c.alignment = LEFT

    if not candidates:
        ws.cell(row=_DATA_START_ROW, column=1,
                value="— No F3 candidates found (E3/E5/E7 users with web/mobile-only usage) —")

    # ── Column widths ─────────────────────────────────────────────────────────
    col_widths = [40, 28, 6, 50, 8, 6, 14, 13, 22, 22, 22]
    for col, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w
    # Price config section uses cols A-C differently — override A/B/C
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 62
