"""
XLA Scorecard sheet — tenant-wide and per-agent experience-level agreement metrics.

Data sources (from the PDF spec):
  viva_reports_cs_session_metrics    → Performance, Quality, Trust, Collaboration
  viva_reports_cs_autonomous_metrics → Performance (autonomous reliability)
  viva_reports_cs_action_metrics     → Quality (action-level success)
  viva_reports_cs_weekly_active_users → Collaboration (WAU trend)
  viva_reports_copilot_adoption      → Collaboration (M365 adoption %)
  kpi_snapshots                      → Collaboration (agent inventory)
"""
from collections import defaultdict
from datetime import datetime, timedelta, timezone

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from src.writers._style import DATA_FONT, LEFT

_CAT_FILL  = PatternFill("solid", fgColor="1F4E79")
_CAT_FONT  = Font(bold=True, color="FFFFFF", size=11)
_SECT_FILL = PatternFill("solid", fgColor="2E75B6")
_SECT_FONT = Font(bold=True, color="FFFFFF", size=11)
_ODD_FILL  = PatternFill("solid", fgColor="FFFFFF")
_EVEN_FILL = PatternFill("solid", fgColor="D6E4F0")


def _pct(num, denom) -> str:
    if denom and num is not None:
        return f"{num / denom * 100:.1f}%"
    return "—"


def _band(ws, row) -> PatternFill:
    return _EVEN_FILL if row % 2 == 0 else _ODD_FILL


def _set(ws, row, col, value, font=None, fill=None, bold=False, align=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = font or (Font(bold=True, size=11) if bold else DATA_FONT)
    if fill:
        c.fill = fill
    c.alignment = align or LEFT
    return c


def write(
    ws: Worksheet,
    session_metrics: list[dict],
    autonomous_metrics: list[dict],
    action_metrics: list[dict],
    wau: list[dict],
    adoption: list[dict],
    kpi_snapshot: dict | None,
    agents: dict[str, dict] | None,
    lookback_days: int = 90,
) -> None:
    ag = agents or {}
    cutoff = (datetime.now(timezone.utc) - timedelta(days=lookback_days)).strftime("%Y-%m-%d")

    # Filter to lookback window
    sess = [r for r in session_metrics     if r.get("metric_date", "") >= cutoff]
    auto = [r for r in autonomous_metrics  if r.get("metric_date", "") >= cutoff]
    acts = [r for r in action_metrics      if r.get("metric_date", "") >= cutoff]

    # ── Aggregate calculations ─────────────────────────────────────────────
    total_sessions = sum(r.get("total_sessions")    or 0 for r in sess)
    engaged        = sum(r.get("engaged_sessions")  or 0 for r in sess)
    resolved       = sum(r.get("resolved_sessions") or 0 for r in sess)
    escalated      = sum(r.get("escalated_sessions") or 0 for r in sess)
    abandoned      = sum(r.get("abandoned_sessions") or 0 for r in sess)
    csat_n         = sum(r.get("csat_responses") or 0 for r in sess)
    csat_pos       = sum((r.get("csat_4") or 0) + (r.get("csat_5") or 0) for r in sess)
    csat_neg       = sum((r.get("csat_1") or 0) + (r.get("csat_2") or 0) for r in sess)
    csat_score_sum = sum(
        (r.get("csat_1") or 0)*1 + (r.get("csat_2") or 0)*2 + (r.get("csat_3") or 0)*3 +
        (r.get("csat_4") or 0)*4 + (r.get("csat_5") or 0)*5
        for r in sess
    )
    ks_engaged_sum  = sum(r.get("ks_engaged")  or 0 for r in sess)
    ks_resolved_sum = sum(r.get("ks_resolved") or 0 for r in sess)

    auto_total   = sum(r.get("total_runs")      or 0 for r in auto)
    auto_success = sum(r.get("successful_runs") or 0 for r in auto)

    act_total   = sum(r.get("total_runs")                or 0 for r in acts)
    act_success = sum(r.get("successful_actions_in_runs") or 0 for r in acts)

    adopt_90d    = [r for r in adoption if r.get("metric_date", "") >= cutoff]
    adopt_ids    = {r.get("person_id") for r in adopt_90d if r.get("person_id")}
    adopt_active = {r.get("person_id") for r in adopt_90d
                    if r.get("person_id") and (r.get("total_copilot_actions") or 0) > 0}
    peak_wau     = max((r.get("active_user_count") or 0 for r in wau), default=0)

    # ── Pre-aggregate monthly + WAU (used in Adoption Health and trend tables) ─
    monthly: dict[str, dict] = defaultdict(lambda: defaultdict(int))
    for r in sess:
        d = r.get("metric_date", "")
        if len(d) >= 7:
            m = d[:7]
            monthly[m]["total"]     += r.get("total_sessions")    or 0
            monthly[m]["engaged"]   += r.get("engaged_sessions")  or 0
            monthly[m]["resolved"]  += r.get("resolved_sessions") or 0
            monthly[m]["escalated"] += r.get("escalated_sessions") or 0
            monthly[m]["abandoned"] += r.get("abandoned_sessions") or 0

    wau_by_week: dict[str, int] = defaultdict(int)
    for r in wau:
        sd = r.get("start_date", "")
        if sd:
            wau_by_week[sd] += r.get("active_user_count") or 0

    # ── Column widths ──────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 52
    ws.freeze_panes = "A3"

    row = 1

    # Title
    ws.merge_cells(f"A{row}:D{row}")
    t = ws.cell(row=row, column=1, value="XLA Scorecard — Agent Experience-Level Agreement Measurements")
    t.font = Font(bold=True, size=14, color="1F4E79")
    t.alignment = LEFT
    ws.row_dimensions[row].height = 26
    row += 1

    ws.merge_cells(f"A{row}:D{row}")
    g = ws.cell(row=row, column=1,
                value=f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}  "
                      f"|  Window: last {lookback_days} days  |  Cutoff: {cutoff}")
    g.font = Font(italic=True, size=10, color="595959")
    g.alignment = LEFT
    row += 2

    def section(label):
        nonlocal row
        ws.merge_cells(f"A{row}:D{row}")
        _set(ws, row, 1, label, font=_CAT_FONT, fill=_CAT_FILL)
        ws.row_dimensions[row].height = 20
        row += 1
        for col, h in enumerate(["Category", "Metric", "Value", "Formula / Source"], 1):
            _set(ws, row, col, h, font=_SECT_FONT, fill=_SECT_FILL)
        row += 1

    def mrow(category, metric, value, formula=""):
        nonlocal row
        fill = _band(ws, row)
        for col, val in enumerate([category, metric, value, formula], 1):
            _set(ws, row, col, val, fill=fill)
        row += 1

    # ── Agent Performance ──────────────────────────────────────────────────
    section("📊  Agent Performance")
    mrow("Agent Performance", "Total Sessions (window)",    total_sessions,
         f"SUM(total_sessions) where metric_date ≥ {cutoff}")
    mrow("Agent Performance", "Engaged Sessions",           engaged,
         "SUM(engaged_sessions)")
    mrow("Agent Performance", "Task Completion Rate",       _pct(resolved, engaged),
         "SUM(resolved_sessions) / SUM(engaged_sessions)")
    mrow("Agent Performance", "Human Escalation Rate",      _pct(escalated, engaged),
         "SUM(escalated_sessions) / SUM(engaged_sessions)")
    mrow("Agent Performance", "Abandonment Rate",           _pct(abandoned, engaged),
         "SUM(abandoned_sessions) / SUM(engaged_sessions)")
    mrow("Agent Performance", "Handoff Success Rate",
         _pct(resolved, (resolved + escalated) or None),
         "SUM(resolved) / (SUM(resolved) + SUM(escalated))")
    if auto_total:
        mrow("Agent Performance", "Autonomous Run Success Rate", _pct(auto_success, auto_total),
             "SUM(successful_runs) / SUM(total_runs) — viva_reports_cs_autonomous_metrics")
        mrow("Agent Performance", "Autonomous Runs (total)",     auto_total,
             "SUM(total_runs) — viva_reports_cs_autonomous_metrics")
    row += 1

    # ── Agent Quality ──────────────────────────────────────────────────────
    section("⭐  Agent Quality (CSAT & Actions)")
    mrow("Agent Quality", "CSAT Responses",        csat_n,
         "SUM(csat_responses)")
    mrow("Agent Quality", "CSAT Positive Rate",    _pct(csat_pos, csat_n),
         "(SUM(csat_4) + SUM(csat_5)) / SUM(csat_responses)")
    mrow("Agent Quality", "CSAT Negative Rate",    _pct(csat_neg, csat_n),
         "(SUM(csat_1) + SUM(csat_2)) / SUM(csat_responses)")
    avg_csat = round(csat_score_sum / csat_n, 2) if csat_n else "—"
    mrow("Agent Quality", "Avg CSAT Score",        avg_csat,
         "Weighted average of csat_1–5 ratings / csat_responses")
    if act_total:
        mrow("Agent Quality", "Action-Level Success Rate", _pct(act_success, act_total),
             "SUM(successful_actions_in_runs) / SUM(total_runs) — viva_reports_cs_action_metrics")
    row += 1

    # ── Trust & Explainability ─────────────────────────────────────────────
    section("🔍  Trust & Explainability")
    mrow("Trust", "KS Coverage (Engaged)",
         _pct(ks_engaged_sum, engaged),
         "SUM(ks_engaged) / SUM(engaged_sessions)")
    mrow("Trust", "KS Coverage (Resolved)",
         _pct(ks_resolved_sum, resolved),
         "SUM(ks_resolved) / SUM(resolved_sessions)")
    row += 1

    # ── Collaboration & Adoption ───────────────────────────────────────────
    section("🤝  Collaboration & Adoption")
    mrow("Collaboration", "M365 Copilot Enabled Users",   len(adopt_ids),
         "COUNT(DISTINCT person_id) — viva_reports_copilot_adoption")
    mrow("Collaboration", "M365 Copilot Active Users",    len(adopt_active),
         "COUNT(DISTINCT person_id WHERE total_copilot_actions > 0)")
    mrow("Collaboration", "M365 Copilot Adoption Rate",
         _pct(len(adopt_active), len(adopt_ids) or None),
         "Active users / Enabled users")
    mrow("Collaboration", "Peak Weekly Active Users (Agents)", peak_wau,
         "MAX(active_user_count) — viva_reports_cs_weekly_active_users")
    if kpi_snapshot:
        mrow("Collaboration", "Total Agents (inventory)",
             kpi_snapshot.get("total_agents") or "—",
             "kpi_snapshots.total_agents")
        mrow("Collaboration", "Production Agents",
             kpi_snapshot.get("production_agents") or "—",
             "kpi_snapshots.production_agents")
        util = kpi_snapshot.get("utilization_rate")
        mrow("Collaboration", "Agent Utilization Rate",
             f"{util:.1f}%" if util else "—",
             "kpi_snapshots.utilization_rate")
        own = kpi_snapshot.get("ownership_pct")
        mrow("Collaboration", "Agent Ownership Coverage",
             f"{own:.1f}%" if own else "—",
             "kpi_snapshots.ownership_pct")
    row += 2

    # ── Readiness vs. Adoption — Supported Indicators ─────────────────────
    section("📉  Readiness vs. Adoption — Supported Indicators")

    month_keys = sorted(monthly.keys())
    wau_keys   = sorted(wau_by_week.keys())

    def _rate(num, denom) -> float:
        return num / denom * 100 if denom else 0.0

    # Workflow Adoption — journey completion trend (first vs. latest month)
    if len(month_keys) >= 2:
        def _res(mk):
            e = monthly[mk].get("engaged") or 0
            return _rate(monthly[mk].get("resolved") or 0, e)
        first_res, latest_res = _res(month_keys[0]), _res(month_keys[-1])
        d = latest_res - first_res
        dir_r = "↑ Growing" if d > 1 else ("↓ Declining" if d < -1 else "→ Stable")
        mrow("Workflow Adoption",
             "Journey Completion Trend (1st → latest month)",
             f"{dir_r}  ({first_res:.1f}% → {latest_res:.1f}%)",
             f"Resolved/Engaged: {month_keys[0]} → {month_keys[-1]}")

    # Quiet Retreat — WAU momentum
    if len(wau_keys) >= 8:
        last_4 = [wau_by_week[w] for w in wau_keys[-4:]]
        prev_4 = [wau_by_week[w] for w in wau_keys[-8:-4]]
        l_avg, p_avg = sum(last_4) / 4, sum(prev_4) / 4
        mom = _rate(l_avg - p_avg, p_avg)
        dir_m = "↑ Growing" if mom > 5 else ("↓ Declining — retreat signal" if mom < -5 else "→ Stable")
        mrow("Quiet Retreat Indicator",
             "WAU Momentum (last 4 wks vs. prior 4 wks)",
             f"{dir_m}  ({mom:+.1f}%)",
             "AVG(last 4 weeks WAU) vs AVG(prior 4 weeks WAU) — viva_reports_cs_weekly_active_users")
    elif len(wau_keys) >= 2:
        fv, lv = wau_by_week[wau_keys[0]], wau_by_week[wau_keys[-1]]
        mom = _rate(lv - fv, fv)
        dir_m = "↑ Growing" if mom > 5 else ("↓ Declining — retreat signal" if mom < -5 else "→ Stable")
        mrow("Quiet Retreat Indicator",
             "WAU Change (first → latest week)",
             f"{dir_m}  ({mom:+.1f}%  |  {fv} → {lv})",
             "First vs. latest week WAU — viva_reports_cs_weekly_active_users")

    # Quiet Retreat — abandonment rate trend
    if len(month_keys) >= 2:
        def _abn(mk):
            e = monthly[mk].get("engaged") or 0
            return _rate(monthly[mk].get("abandoned") or 0, e)
        first_abn, latest_abn = _abn(month_keys[0]), _abn(month_keys[-1])
        d = latest_abn - first_abn
        dir_a = "↑ Rising — retreat signal" if d > 1 else ("↓ Falling" if d < -1 else "→ Stable")
        mrow("Quiet Retreat Indicator",
             "Abandonment Rate Trend (1st → latest month)",
             f"{dir_a}  ({first_abn:.1f}% → {latest_abn:.1f}%)",
             f"Abandoned/Engaged: {month_keys[0]} → {month_keys[-1]}")

    # Hypercare Demand — escalation rate trend
    if len(month_keys) >= 2:
        def _esc(mk):
            e = monthly[mk].get("engaged") or 0
            return _rate(monthly[mk].get("escalated") or 0, e)
        first_esc, latest_esc = _esc(month_keys[0]), _esc(month_keys[-1])
        d = latest_esc - first_esc
        dir_e = "↑ Rising — demand signal" if d > 1 else ("↓ Falling" if d < -1 else "→ Stable")
        mrow("Hypercare Demand",
             "Escalation Rate Trend (1st → latest month)",
             f"{dir_e}  ({first_esc:.1f}% → {latest_esc:.1f}%)",
             f"Escalated/Engaged: {month_keys[0]} → {month_keys[-1]}")

    # Operating Model Durability — resolution rate consistency
    if len(month_keys) >= 2:
        res_rates = []
        for mk in month_keys:
            e = monthly[mk].get("engaged") or 0
            r = monthly[mk].get("resolved") or 0
            if e:
                res_rates.append(r / e * 100)
        if res_rates:
            mn, mx = min(res_rates), max(res_rates)
            spread = mx - mn
            stability = "Stable" if spread < 5 else ("Moderate" if spread < 15 else "Variable")
            mrow("Operating Model Durability",
                 "Resolution Rate Consistency (monthly spread)",
                 f"{stability}  ({mn:.1f}% – {mx:.1f}%,  ±{spread:.1f}pp spread)",
                 "Min/max monthly Resolved/Engaged rate across window")

    # System Reliance — agent WAU vs. M365 Copilot enabled users
    if adopt_ids and wau_keys:
        latest_wau_val = wau_by_week[wau_keys[-1]]
        reliance_pct   = _rate(latest_wau_val, len(adopt_ids))
        mrow("System Reliance",
             "Agent WAU vs. M365 Copilot Enabled Users (latest week)",
             f"{reliance_pct:.1f}%  ({latest_wau_val:,} WAU  /  {len(adopt_ids):,} enabled)",
             "Latest week agent WAU ÷ COUNT(DISTINCT person_id) — viva_reports_copilot_adoption")

    row += 1

    # ── Monthly Session Trend ──────────────────────────────────────────────
    ws.merge_cells(f"A{row}:F{row}")
    _set(ws, row, 1, "📅  Monthly Session Trend", font=_CAT_FONT, fill=_CAT_FILL)
    ws.row_dimensions[row].height = 20
    row += 1
    for col, h in enumerate(["Month", "Total Sessions", "Engaged", "Resolved",
                              "Escalated", "Abandoned"], 1):
        _set(ws, row, col, h, font=_SECT_FONT, fill=_SECT_FILL)
    for ltr, w in [("E", 12), ("F", 12)]:
        ws.column_dimensions[ltr].width = w
    row += 1

    for month in sorted(monthly):
        m = monthly[month]
        fill = _band(ws, row)
        for col, val in enumerate([month, m["total"], m["engaged"],
                                   m["resolved"], m["escalated"], m["abandoned"]], 1):
            _set(ws, row, col, val, fill=fill)
        row += 1
    row += 2

    # ── WAU Trend ──────────────────────────────────────────────────────────
    ws.merge_cells(f"A{row}:B{row}")
    _set(ws, row, 1, "📈  Weekly Active Users Trend (all time)", font=_CAT_FONT, fill=_CAT_FILL)
    ws.row_dimensions[row].height = 20
    row += 1
    for col, h in enumerate(["Week Start", "Active Users (all agents)"], 1):
        _set(ws, row, col, h, font=_SECT_FONT, fill=_SECT_FILL)
    ws.column_dimensions["B"].width = max(ws.column_dimensions["B"].width, 26)
    row += 1

    # Aggregate WAU across all agents per week
    for week in sorted(wau_by_week):
        fill = _band(ws, row)
        _set(ws, row, 1, week,              fill=fill)
        _set(ws, row, 2, wau_by_week[week], fill=fill)
        row += 1
    row += 2

    # ── Per-Agent XLA Summary ──────────────────────────────────────────────
    ws.merge_cells(f"A{row}:H{row}")
    _set(ws, row, 1, "🤖  Per-Agent XLA Summary (window)", font=_CAT_FONT, fill=_CAT_FILL)
    ws.row_dimensions[row].height = 20
    row += 1
    agent_hdrs = ["Agent Name", "Sessions", "Completion %", "Escalation %",
                  "Abandonment %", "CSAT Responses", "Avg CSAT", "KS Coverage (Engaged)"]
    for col, h in enumerate(agent_hdrs, 1):
        _set(ws, row, col, h, font=_SECT_FONT, fill=_SECT_FILL)
    for ltr, w in [("E", 15), ("F", 16), ("G", 12), ("H", 22)]:
        ws.column_dimensions[ltr].width = w
    row += 1

    # Aggregate per agent_id
    by_agent: dict[str, dict] = defaultdict(lambda: defaultdict(int))
    for r in sess:
        aid = r.get("agent_id", "")
        by_agent[aid]["total"]     += r.get("total_sessions")    or 0
        by_agent[aid]["engaged"]   += r.get("engaged_sessions")  or 0
        by_agent[aid]["resolved"]  += r.get("resolved_sessions") or 0
        by_agent[aid]["escalated"] += r.get("escalated_sessions") or 0
        by_agent[aid]["abandoned"] += r.get("abandoned_sessions") or 0
        by_agent[aid]["csat_n"]    += r.get("csat_responses")    or 0
        by_agent[aid]["csat_s"]    += (
            (r.get("csat_1") or 0)*1 + (r.get("csat_2") or 0)*2 +
            (r.get("csat_3") or 0)*3 + (r.get("csat_4") or 0)*4 +
            (r.get("csat_5") or 0)*5
        )
        by_agent[aid]["ks_engaged"] += r.get("ks_engaged") or 0

    for aid, m in sorted(by_agent.items(),
                         key=lambda x: x[1]["total"], reverse=True):
        name = ag.get(aid, {}).get("agent_name", aid)
        avg  = round(m["csat_s"] / m["csat_n"], 2) if m["csat_n"] else "—"
        fill = _band(ws, row)
        for col, val in enumerate([
            name,
            m["total"],
            _pct(m["resolved"],  m["engaged"] or None),
            _pct(m["escalated"], m["engaged"] or None),
            _pct(m["abandoned"], m["engaged"] or None),
            m["csat_n"],
            avg,
            _pct(m["ks_engaged"], m["engaged"] or None),
        ], 1):
            _set(ws, row, col, val, fill=fill)
        row += 1
