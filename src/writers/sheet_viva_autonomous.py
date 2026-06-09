from openpyxl.worksheet.worksheet import Worksheet

from src.writers._style import apply_row_style, autofit_columns, write_headers

HEADERS = [
    "Agent ID", "Agent Name", "Metric Date",
    "Total Runs", "Successful", "Failed", "Success %",
    "Total Duration (s)", "Avg Duration (s)",
    "Runs w/ Knowledge", "Runs w/ Actions", "Runs (No Ops)",
]


def _pct(num, denom):
    if denom and denom > 0 and num is not None:
        return round(num / denom * 100, 1)
    return None


def _avg(total, runs):
    if runs and runs > 0 and total is not None:
        return round(total / runs, 1)
    return None


def write(ws: Worksheet, rows: list[dict], agents: dict[str, dict] | None = None) -> None:
    write_headers(ws, HEADERS)
    ag = agents or {}

    if not rows:
        ws.cell(row=2, column=1, value="— No autonomous metrics imported —")
    else:
        for i, r in enumerate(rows, start=2):
            aid = r.get('agent_id', '')
            total = r.get('total_runs')
            succ  = r.get('successful_runs')
            dur   = r.get('total_duration')
            ws.cell(row=i, column=1,  value=aid)
            ws.cell(row=i, column=2,  value=ag.get(aid, {}).get('agent_name', ''))
            ws.cell(row=i, column=3,  value=r.get('metric_date', ''))
            ws.cell(row=i, column=4,  value=total)
            ws.cell(row=i, column=5,  value=succ)
            ws.cell(row=i, column=6,  value=r.get('failed_runs'))
            ws.cell(row=i, column=7,  value=_pct(succ, total))
            ws.cell(row=i, column=8,  value=dur)
            ws.cell(row=i, column=9,  value=_avg(dur, total))
            ws.cell(row=i, column=10, value=r.get('ks_successful'))
            ws.cell(row=i, column=11, value=r.get('actions_successful'))
            ws.cell(row=i, column=12, value=r.get('no_op_successful'))
            apply_row_style(ws, i, len(HEADERS))

    autofit_columns(ws, HEADERS)
