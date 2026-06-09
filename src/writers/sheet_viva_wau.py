from openpyxl.worksheet.worksheet import Worksheet

from src.writers._style import apply_row_style, autofit_columns, write_headers

HEADERS = ["Agent ID", "Agent Name", "Week Start", "Active Users"]


def write(ws: Worksheet, rows: list[dict], agents: dict[str, dict] | None = None) -> None:
    write_headers(ws, HEADERS)
    ag = agents or {}

    if not rows:
        ws.cell(row=2, column=1, value="— No weekly active user data imported —")
    else:
        for i, r in enumerate(rows, start=2):
            aid = r.get('agent_id', '')
            ws.cell(row=i, column=1, value=aid)
            ws.cell(row=i, column=2, value=ag.get(aid, {}).get('agent_name', ''))
            ws.cell(row=i, column=3, value=r.get('start_date', ''))
            ws.cell(row=i, column=4, value=r.get('active_user_count'))
            apply_row_style(ws, i, len(HEADERS))

    autofit_columns(ws, HEADERS)
