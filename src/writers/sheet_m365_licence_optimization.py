from collections import defaultdict
from datetime import datetime, timezone

from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from src.writers._style import HEADER_FILL, HEADER_FONT, LEFT

_WARN_FILL  = PatternFill("solid", fgColor="FFF2CC")   # amber — watch
_ALERT_FILL = PatternFill("solid", fgColor="FFE0E0")   # red   — action needed
_GOOD_FILL  = PatternFill("solid", fgColor="E2EFDA")   # green — healthy


def write(
    ws: Worksheet,
    billing_licences: list[dict],
    services_counts: list[dict],
    active_users_detail: list[dict],
    activations_users: list[dict],
    proplus_counts: list[dict],
) -> None:

    # ── Section 1: License inventory ─────────────────────────────────────────
    total_all    = sum(r.get("total_licenses")    or 0 for r in billing_licences)
    assigned_all = sum(r.get("assigned_licenses") or 0 for r in billing_licences)
    expired_all  = sum(r.get("expired_licenses")  or 0 for r in billing_licences)
    unassigned_all = max(0, total_all - assigned_all)
    assignment_rate = round(assigned_all / total_all * 100, 1) if total_all else 0.0

    # Per-product unassigned — only SKUs with at least 1 unassigned
    product_unassigned: list[tuple[str, int, int, int]] = []
    for r in billing_licences:
        total    = r.get("total_licenses")    or 0
        assigned = r.get("assigned_licenses") or 0
        expired  = r.get("expired_licenses")  or 0
        unused   = max(0, total - assigned)
        if unused > 0:
            product_unassigned.append((r.get("product_title", "Unknown"), total, assigned, unused))
    product_unassigned.sort(key=lambda x: x[3], reverse=True)

    # ── Section 2: Service adoption (most recent snapshot) ───────────────────
    snap = sorted(services_counts, key=lambda r: r.get("report_refresh_date", ""), reverse=True)
    snap = snap[0] if snap else {}

    def _util(active_key: str, inactive_key: str) -> tuple[int, int, float]:
        active   = snap.get(active_key)   or 0
        inactive = snap.get(inactive_key) or 0
        total    = active + inactive
        pct      = round(active / total * 100, 1) if total else 0.0
        return active, inactive, pct

    exc_a, exc_i, exc_pct   = _util("exchange_active",  "exchange_inactive")
    od_a,  od_i,  od_pct    = _util("onedrive_active",  "onedrive_inactive")
    sp_a,  sp_i,  sp_pct    = _util("sharepoint_active","sharepoint_inactive")
    t_a,   t_i,   t_pct     = _util("teams_active",     "teams_inactive")
    ym_a,  ym_i,  ym_pct    = _util("yammer_active",    "yammer_inactive")
    o365_a,o365_i,o365_pct  = _util("office365_active", "office365_inactive")

    snap_date = snap.get("report_refresh_date", "")

    # ── Section 3: Inactive licensed users (per-user detail) ─────────────────
    def _inactive_count(has_key: str, activity_key: str) -> tuple[int, int]:
        licensed = [r for r in active_users_detail if r.get(has_key) and not r.get("is_deleted")]
        inactive = [r for r in licensed if not r.get(activity_key)]
        return len(licensed), len(inactive)

    exc_lic, exc_inactive     = _inactive_count("has_exchange",  "exchange_last_activity")
    od_lic,  od_inactive      = _inactive_count("has_onedrive",  "onedrive_last_activity")
    sp_lic,  sp_inactive      = _inactive_count("has_sharepoint","sharepoint_last_activity")
    t_lic,   t_inactive       = _inactive_count("has_teams",     "teams_last_activity")
    ym_lic,  ym_inactive      = _inactive_count("has_yammer",    "yammer_last_activity")

    def _inactive_pct(lic: int, inactive: int) -> float:
        return round(inactive / lic * 100, 1) if lic else 0.0

    # ── Section 4: Never-activated products ───────────────────────────────────
    never_by_product: dict[str, int] = defaultdict(int)
    for r in activations_users:
        if not r.get("last_activated_date"):
            never_by_product[r.get("product_type", "Unknown")] += 1
    never_sorted = sorted(never_by_product.items(), key=lambda x: x[1], reverse=True)

    # ── Section 5: ProPlus app adoption (most recent snapshot) ───────────────
    pp_snap = sorted(proplus_counts, key=lambda r: r.get("report_date", ""), reverse=True)
    pp = pp_snap[0] if pp_snap else {}

    # ── Helpers ───────────────────────────────────────────────────────────────
    def _fmt(v) -> str:
        if v is None:
            return "—"
        if isinstance(v, float):
            return f"{v:,.1f}"
        if isinstance(v, int):
            return f"{v:,}"
        return str(v)

    # ── Build row list ────────────────────────────────────────────────────────
    rows: list[tuple] = [
        ("Report generated", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")),
        ("Snapshot date",    snap_date or "—"),
    ]

    # ── Section 1: Inventory overview ─────────────────────────────────────────
    rows += [
        (None, None),
        ("── License Inventory ────────────────────────────────", None),
        ("Total licenses (all products)", _fmt(total_all)),
        ("Total assigned",               _fmt(assigned_all)),
        ("Total unassigned",             _fmt(unassigned_all)),
        ("Total expired",                _fmt(expired_all)),
        ("Overall assignment rate",      f"{assignment_rate}%"),
        ("Products with unassigned seats", _fmt(len(product_unassigned))),
    ]

    # ── Section 2: Top unassigned SKUs ────────────────────────────────────────
    rows += [
        (None, None),
        ("── Top Unassigned License SKUs ──────────────────────", None),
    ]
    for title, total, assigned, unused in product_unassigned[:15]:
        pct_used = round(assigned / total * 100, 1) if total else 0.0
        rows.append((f"  {title}", f"{_fmt(unused)} unassigned  ({pct_used}% assigned of {_fmt(total)})"))

    # ── Section 3: Service adoption ───────────────────────────────────────────
    rows += [
        (None, None),
        ("── Service Active User Rate ─────────────────────────", None),
        ("  Exchange",   f"{exc_pct}%  ({_fmt(exc_a)} active  /  {_fmt(exc_a + exc_i)} total)"),
        ("  OneDrive",   f"{od_pct}%  ({_fmt(od_a)} active  /  {_fmt(od_a + od_i)} total)"),
        ("  SharePoint", f"{sp_pct}%  ({_fmt(sp_a)} active  /  {_fmt(sp_a + sp_i)} total)"),
        ("  Teams",      f"{t_pct}%  ({_fmt(t_a)} active  /  {_fmt(t_a + t_i)} total)"),
        ("  Yammer",     f"{ym_pct}%  ({_fmt(ym_a)} active  /  {_fmt(ym_a + ym_i)} total)"),
        ("  Office 365", f"{o365_pct}%  ({_fmt(o365_a)} active  /  {_fmt(o365_a + o365_i)} total)"),
    ]

    # ── Section 4: Inactive licensed users ────────────────────────────────────
    rows += [
        (None, None),
        ("── Inactive Licensed Users (no activity in period) ──", None),
        ("  Exchange  — licensed / inactive",
         f"{_fmt(exc_lic)} licensed  →  {_fmt(exc_inactive)} inactive  ({_inactive_pct(exc_lic, exc_inactive)}%)"),
        ("  OneDrive  — licensed / inactive",
         f"{_fmt(od_lic)} licensed  →  {_fmt(od_inactive)} inactive  ({_inactive_pct(od_lic, od_inactive)}%)"),
        ("  SharePoint — licensed / inactive",
         f"{_fmt(sp_lic)} licensed  →  {_fmt(sp_inactive)} inactive  ({_inactive_pct(sp_lic, sp_inactive)}%)"),
        ("  Teams     — licensed / inactive",
         f"{_fmt(t_lic)} licensed  →  {_fmt(t_inactive)} inactive  ({_inactive_pct(t_lic, t_inactive)}%)"),
        ("  Yammer    — licensed / inactive",
         f"{_fmt(ym_lic)} licensed  →  {_fmt(ym_inactive)} inactive  ({_inactive_pct(ym_lic, ym_inactive)}%)"),
    ]

    # ── Section 5: Never-activated products ───────────────────────────────────
    if never_sorted:
        rows += [
            (None, None),
            ("── Never-Activated Product Assignments ──────────────", None),
        ]
        for product, count in never_sorted[:10]:
            rows.append((f"  {product}", f"{_fmt(count)} users never activated"))

    # ── Section 6: ProPlus app adoption ───────────────────────────────────────
    if pp:
        pp_date = pp.get("report_date", "")
        rows += [
            (None, None),
            ("── Microsoft 365 Apps Active Users ──────────────────", None),
            ("  (report date)", pp_date),
            ("  Outlook",    _fmt(pp.get("outlook"))),
            ("  Word",       _fmt(pp.get("word"))),
            ("  Excel",      _fmt(pp.get("excel"))),
            ("  PowerPoint", _fmt(pp.get("powerpoint"))),
            ("  OneNote",    _fmt(pp.get("onenote"))),
            ("  Teams",      _fmt(pp.get("teams"))),
        ]

    # ── Write headers ─────────────────────────────────────────────────────────
    for col, header in enumerate(["Metric", "Value"], 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = LEFT

    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 56
    ws.column_dimensions["B"].width = 52

    # ── Write rows ────────────────────────────────────────────────────────────
    for row_idx, (metric, value) in enumerate(rows, 2):
        a = ws.cell(row=row_idx, column=1, value=metric)
        b = ws.cell(row=row_idx, column=2, value=value)
        a.alignment = LEFT
        b.alignment = LEFT

        is_section = bool(metric and metric.startswith("──"))
        if is_section:
            a.font = Font(bold=True, size=11, color="1F4E79")
            b.font = Font(size=11)
        else:
            a.font = Font(size=11)
            b.font = Font(size=11)

        val_str = str(value or "")

        # Assignment rate
        if metric == "Overall assignment rate":
            try:
                pct_v = float(val_str.replace("%", ""))
                if pct_v >= 85:
                    a.fill = _GOOD_FILL;  b.fill = _GOOD_FILL
                elif pct_v >= 70:
                    a.fill = _WARN_FILL;  b.fill = _WARN_FILL
                else:
                    a.fill = _ALERT_FILL; b.fill = _ALERT_FILL
            except (ValueError, TypeError):
                pass

        # Service utilization lines
        elif metric and metric.startswith("  ") and "active  /" in val_str and "%" in val_str:
            try:
                pct_v = float(val_str.split("%")[0].strip())
                if pct_v >= 70:
                    a.fill = _GOOD_FILL;  b.fill = _GOOD_FILL
                elif pct_v >= 50:
                    a.fill = _WARN_FILL;  b.fill = _WARN_FILL
                else:
                    a.fill = _ALERT_FILL; b.fill = _ALERT_FILL
            except (ValueError, TypeError):
                pass

        # Inactive licensed user lines
        elif "licensed  →" in val_str and "inactive" in val_str:
            try:
                pct_v = float(val_str.split("(")[1].split("%")[0].strip())
                if pct_v == 0:
                    a.fill = _GOOD_FILL;  b.fill = _GOOD_FILL
                elif pct_v < 20:
                    a.fill = _WARN_FILL;  b.fill = _WARN_FILL
                else:
                    a.fill = _ALERT_FILL; b.fill = _ALERT_FILL
            except (ValueError, TypeError, IndexError):
                pass
