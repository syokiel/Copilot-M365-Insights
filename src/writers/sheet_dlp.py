from openpyxl.worksheet.worksheet import Worksheet

from src.writers._style import autofit_columns, write_headers

HEADERS = [
    "Policy ID",
    "Display Name",
    "Environment Type",
    "Created By",
    "Created Date",
    "Modified Date",
    "Enforcement Mode",
    "Blocked Connectors",
    "Business Connectors",
    "Non-Business Connectors",
]


def write(ws: Worksheet, policies: list[dict]) -> None:
    write_headers(ws, HEADERS)

    if not policies:
        ws.cell(row=2, column=1, value="— Requires Power Platform Admin role on the sync service principal —")
    else:
        for i, p in enumerate(policies, start=2):
            ws.cell(row=i, column=1, value=p.get("policy_id", ""))
            ws.cell(row=i, column=2, value=p.get("display_name", ""))
            ws.cell(row=i, column=3, value=p.get("environment_type", ""))
            ws.cell(row=i, column=4, value=p.get("created_by", ""))
            ws.cell(row=i, column=5, value=p.get("created_at", ""))
            ws.cell(row=i, column=6, value=p.get("modified_at", ""))
            ws.cell(row=i, column=7, value=p.get("enforcement_mode", ""))
            ws.cell(row=i, column=8, value=p.get("blocked_connectors", ""))
            ws.cell(row=i, column=9, value=p.get("business_connectors", ""))
            ws.cell(row=i, column=10, value=p.get("non_business_connectors", ""))

    autofit_columns(ws, HEADERS)
