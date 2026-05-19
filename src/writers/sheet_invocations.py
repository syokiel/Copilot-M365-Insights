from collections import defaultdict

from openpyxl.worksheet.worksheet import Worksheet

from src.writers._style import apply_row_style, autofit_columns, write_headers

HEADERS = [
    "Conversation ID",
    "Session ID",
    "User ID",
    "Channel",
    "Design Mode",
    "First Message",
    "Last Message",
    "Duration (min)",
    "Messages Received",
    "Messages Sent",
    "Topics",
    "Connector Calls",
]


def _build_conversations(events: list[dict], connector_calls: list[dict]) -> list[dict]:
    by_conv: dict[str, dict] = defaultdict(lambda: {
        "session_id": "", "user_id": "", "channel_id": "",
        "design_mode": False, "timestamps": [],
        "received": 0, "sent": 0, "topics": set(), "connector_calls": 0,
    })

    for e in events:
        cid = e.get("ConversationId") or ""
        if not cid:
            continue
        c = by_conv[cid]
        c["session_id"] = c["session_id"] or e.get("SessionId", "")
        c["user_id"] = c["user_id"] or e.get("UserId", "")
        c["channel_id"] = c["channel_id"] or e.get("ChannelId", "")
        c["design_mode"] = bool(e.get("DesignMode"))
        if e.get("Timestamp"):
            c["timestamps"].append(e["Timestamp"])
        name = e.get("EventName", "")
        if name == "BotMessageReceived":
            c["received"] += 1
        elif name == "BotMessageSend":
            c["sent"] += 1
        topic = e.get("TopicName", "")
        if topic:
            c["topics"].add(topic)

    for cc in connector_calls:
        cid = cc.get("ConversationId") or ""
        if cid in by_conv:
            by_conv[cid]["connector_calls"] += 1

    return [{"conversation_id": cid, **data} for cid, data in by_conv.items()]


def write(ws: Worksheet, events: list[dict], connector_calls: list[dict]) -> None:
    conversations = _build_conversations(events, connector_calls)
    conversations.sort(key=lambda c: min(c["timestamps"]) if c["timestamps"] else "", reverse=True)

    write_headers(ws, HEADERS)

    for row_idx, c in enumerate(conversations, 2):
        ts = c["timestamps"]
        first = min(ts) if ts else None
        last = max(ts) if ts else None
        duration = ""
        if first and last:
            delta = last - first
            duration = round(delta.total_seconds() / 60, 1)

        ws.cell(row=row_idx, column=1, value=c["conversation_id"])
        ws.cell(row=row_idx, column=2, value=c["session_id"])
        ws.cell(row=row_idx, column=3, value=c["user_id"])
        ws.cell(row=row_idx, column=4, value=c["channel_id"])
        ws.cell(row=row_idx, column=5, value="Yes" if c["design_mode"] else "No")
        ws.cell(row=row_idx, column=6, value=str(first)[:19] if first else "")
        ws.cell(row=row_idx, column=7, value=str(last)[:19] if last else "")
        ws.cell(row=row_idx, column=8, value=duration)
        ws.cell(row=row_idx, column=9, value=c["received"])
        ws.cell(row=row_idx, column=10, value=c["sent"])
        ws.cell(row=row_idx, column=11, value=", ".join(sorted(c["topics"])))
        ws.cell(row=row_idx, column=12, value=c["connector_calls"])
        apply_row_style(ws, row_idx, len(HEADERS))

    autofit_columns(ws, HEADERS)
