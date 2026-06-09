"""
Two sections on one sheet:
  - Summary row (latest refresh date counts)
  - Daily trend table
"""
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet

from src.writers._style import (
    HEADER_FILL, HEADER_FONT, LEFT, apply_row_style, autofit_columns, write_headers,
)

_APP_COLS = [
    ("chat_active",           "Copilot Chat"),
    ("teams_active",          "Teams Chat"),
    ("teams_meetings_active", "Teams Meetings"),
    ("word_active",           "Word"),
    ("excel_active",          "Excel"),
    ("powerpoint_active",     "PowerPoint"),
    ("outlook_active",        "Outlook"),
    ("onenote_active",        "OneNote"),
    ("loop_active",           "Loop"),
]

SUMMARY_HEADERS = ["Metric", "Count"]
TREND_HEADERS   = ["Date", "Active Users"] + [label for _, label in _APP_COLS]


def write(
    ws: Worksheet,
    summary_rows: list[dict],
    trend_rows: list[dict],
) -> None:
    row = 1

    # ── Summary section ───────────────────────────────────────────────────────
    title = ws.cell(row=row, column=1, value="── Copilot User Count Summary ──────────────────")
    title.font = Font(bold=True, size=11, color="1F4E79")
    row += 1

    if summary_rows:
        latest = summary_rows[0]  # already sorted DESC by refresh date
        write_headers(ws, SUMMARY_HEADERS, start_row=row)
        row += 1
        summary_items = [
            ("Report Refresh Date",  latest.get("report_refresh_date")),
            ("Report Period",         latest.get("report_period")),
            ("Enabled Users",         latest.get("enabled_users")),
            ("Active Users",          latest.get("active_users")),
            ("Copilot Chat",          latest.get("chat_active")),
            ("Teams Chat",            latest.get("teams_active")),
            ("Teams Meetings",        latest.get("teams_meetings_active")),
            ("Word",                  latest.get("word_active")),
            ("Excel",                 latest.get("excel_active")),
            ("PowerPoint",            latest.get("powerpoint_active")),
            ("Outlook",               latest.get("outlook_active")),
            ("OneNote",               latest.get("onenote_active")),
            ("Loop",                  latest.get("loop_active")),
            ("Windows (any Copilot)", latest.get("windows_active")),
            ("Web (any Copilot)",     latest.get("web_active")),
            ("Mobile (any Copilot)",  latest.get("mobile_active")),
        ]
        for metric, value in summary_items:
            ws.cell(row=row, column=1, value=metric).font = Font(size=11)
            ws.cell(row=row, column=2, value=value).font = Font(size=11)
            ws.cell(row=row, column=1).alignment = LEFT
            ws.cell(row=row, column=2).alignment = LEFT
            row += 1
    else:
        ws.cell(row=row, column=1, value="— No summary data available —")
        row += 1

    row += 1  # spacer

    # ── Trend table ───────────────────────────────────────────────────────────
    title2 = ws.cell(row=row, column=1, value="── Daily Active User Trend ─────────────────────")
    title2.font = Font(bold=True, size=11, color="1F4E79")
    row += 1

    write_headers(ws, TREND_HEADERS, start_row=row)
    row += 1

    if trend_rows:
        for r in trend_rows:
            ws.cell(row=row, column=1, value=r.get("report_date", ""))
            ws.cell(row=row, column=2, value=r.get("active_users"))
            for col_idx, (field, _) in enumerate(_APP_COLS, start=3):
                ws.cell(row=row, column=col_idx, value=r.get(field))
            apply_row_style(ws, row, len(TREND_HEADERS))
            row += 1
    else:
        ws.cell(row=row, column=1, value="— No trend data available —")

    ws.column_dimensions["A"].width = 28
    for col in "BCDEFGHIJK":
        ws.column_dimensions[col].width = 16
