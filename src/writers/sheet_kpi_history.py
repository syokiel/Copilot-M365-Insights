from collections import defaultdict

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from src.writers._style import LEFT

# ---------------------------------------------------------------------------
# Per-section colour palette
# Each entry: (label, col_start, col_end, light_fill_hex, dark_fill_hex)
# col_start / col_end are 1-based and are filled in dynamically at write time
# because the LLM section width depends on how many models are in the data.
# ---------------------------------------------------------------------------

_PALETTE = [
    # label               light_fill   dark_fill
    ("",                  "F2F2F2",    "595959"),   # meta / date
    ("M365 Copilot",      "DEEAF1",    "2E75B6"),   # blue
    ("Workloads",         "E2EFDA",    "538135"),   # green
    ("Agent Adoption",    "FFF2CC",    "BF8F00"),   # amber
    ("Agent Inventory",   "FCE4D6",    "C55A11"),   # orange
    ("Environments",      "EAD7F5",    "7030A0"),   # purple
    ("Entitlement",       "D6E4F0",    "1F4E79"),   # navy
    ("Capacity Burn",     "FFE0E0",    "C00000"),   # red
    ("Top LLM Models",    "D9EAD3",    "274E13"),   # dark green
]

_BASE_HEADERS = [
    # meta (2)
    "Snapshot Date", "Lookback Days",
    # M365 Copilot (8)
    "Total Licenses", "Enabled Users", "Active Users",
    "Activation Rate", "Adoption Rate", "Power Users",
    "Total Prompts", "Avg Prompts / User",
    # Workloads (8)
    "Copilot Chat Prompts", "Teams Prompts", "Outlook Prompts",
    "Excel Prompts", "Word Prompts", "PowerPoint Prompts",
    "OneNote Prompts", "Loop Prompts",
    # Agent Adoption (2)
    "Agent Adopters", "Agent Adoption %",
    # Agent Inventory (8)
    "Total Agents", "Active Agents", "Utilization Rate",
    "Production Agents", "Non-Prod Agents",
    "Agents with Owner", "Ownership %", "Total Conversations",
    # Environments (6)
    "Env: Default", "Env: Developer", "Env: Teams",
    "Env: Production", "Env: Sandbox", "Env: Trial",
    # Entitlement (5)
    "Entitled Credits", "Prepaid Consumed", "PAYG Consumed",
    "Remaining Credits", "% Entitlement Used",
    # Capacity Burn (3)
    "Total Capacity", "Avg Daily Burn", "Days Remaining",
    # Top LLM Models — dynamic: appended per-model at write time
]

# Column ranges for fixed sections (1-based)
_FIXED_SECTIONS = [
    (0,  1,  2),   # meta
    (1,  3,  10),  # M365 Copilot
    (2,  11, 18),  # Workloads
    (3,  19, 20),  # Agent Adoption
    (4,  21, 28),  # Agent Inventory
    (5,  29, 34),  # Environments
    (6,  35, 39),  # Entitlement
    (7,  40, 42),  # Capacity Burn
    # index 8 (Top LLM Models) is computed dynamically
]

_PCT_COLS = {6, 7, 20, 23, 27, 39}   # cols whose values are fractions stored as %/100


def write(
    ws: Worksheet,
    snapshots: list[dict],
    entitlement: list[dict] | None = None,
    per_agent: list[dict] | None = None,
    capacity: list[dict] | None = None,
) -> None:

    entitlement = entitlement or []
    per_agent   = per_agent   or []
    capacity    = capacity    or []

    # ── Tokenomics aggregates (current import — same value on every snapshot row) ──
    total_entitled = sum(r.get("entitled_quantity")         or 0 for r in entitlement)
    total_prepaid  = sum(r.get("prepaid_consumed_quantity") or 0 for r in entitlement)
    total_payg     = sum(r.get("payg_consumed_quantity")    or 0 for r in entitlement)
    total_consumed = total_prepaid + total_payg
    remaining      = max(0.0, total_entitled - total_prepaid)
    pct_used       = round(total_consumed / total_entitled * 100, 1) if total_entitled else None

    daily: dict[str, float] = defaultdict(float)
    billable = unbillable = 0.0
    for r in capacity:
        qty = r.get("consumed_quantity") or 0
        d   = r.get("consumption_date", "")
        if d:
            daily[d[:10]] += qty
        if r.get("is_billable"):
            billable += qty
        else:
            unbillable += qty
    total_cap  = billable + unbillable
    avg_daily  = round(total_cap / len(daily), 1) if daily else None
    days_left  = round(remaining / avg_daily) if avg_daily and remaining else None

    # Top 5 LLM models by total credits
    model_totals: dict[str, float] = defaultdict(float)
    for r in per_agent:
        m = r.get("llm_model") or "Unknown"
        model_totals[m] += (r.get("billed_credit") or 0) + (r.get("non_billed_credit") or 0)
    top_models = sorted(model_totals.items(), key=lambda x: x[1], reverse=True)[:5]
    llm_headers = [m for m, _ in top_models]

    # ── Build full header list and section definitions ──────────────────────
    headers = _BASE_HEADERS + llm_headers

    # Dynamic section for LLM models
    llm_start = 43
    llm_end   = llm_start + len(llm_headers) - 1 if llm_headers else llm_start

    sections_with_cols = _FIXED_SECTIONS[:]
    if llm_headers:
        sections_with_cols.append((8, llm_start, llm_end))

    def _make_fill(hex_color: str) -> PatternFill:
        return PatternFill("solid", fgColor=hex_color)

    def _make_font(hex_color: str, bold: bool = False) -> Font:
        return Font(bold=bold, color=hex_color, size=10)

    center = Alignment(horizontal="center", vertical="center")
    left   = LEFT

    # ── Row 1: section group labels ─────────────────────────────────────────
    for palette_idx, col_start, col_end in sections_with_cols:
        label, light_hex, dark_hex = _PALETTE[palette_idx]
        if not label:
            continue
        cell = ws.cell(row=1, column=col_start, value=label)
        cell.fill      = _make_fill(light_hex)
        cell.font      = Font(bold=True, size=10, color=dark_hex)
        cell.alignment = center
        if col_end > col_start:
            ws.merge_cells(
                start_row=1, start_column=col_start,
                end_row=1,   end_column=col_end,
            )

    # ── Row 2: column headers (per-section dark fill, white text) ───────────
    for palette_idx, col_start, col_end in sections_with_cols:
        _label, _light, dark_hex = _PALETTE[palette_idx]
        header_fill = _make_fill(dark_hex)
        header_font = Font(bold=True, color="FFFFFF", size=10)
        for col in range(col_start, col_end + 1):
            h_idx = col - 1
            if h_idx >= len(headers):
                break
            cell = ws.cell(row=2, column=col, value=headers[h_idx])
            cell.fill      = header_fill
            cell.font      = header_font
            cell.alignment = center

    ws.freeze_panes         = "A3"
    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 20

    # ── Data rows ────────────────────────────────────────────────────────────
    if not snapshots:
        ws.cell(row=3, column=1, value="— No KPI snapshots yet — run sync to create the first one —")
    else:
        row_fill_even = PatternFill("solid", fgColor="EBF3FB")
        row_fill_odd  = PatternFill("solid", fgColor="FFFFFF")
        data_font     = Font(size=10)

        def _pct(v):
            return (v / 100) if v is not None else None

        for row_idx, snap in enumerate(snapshots, start=3):
            vals = [
                (snap.get("snapshot_date") or "")[:19],
                snap.get("lookback_days"),
                # M365 Copilot
                snap.get("total_licenses"),
                snap.get("enabled_users"),
                snap.get("active_users"),
                _pct(snap.get("activation_rate")),
                _pct(snap.get("adoption_rate")),
                snap.get("power_users"),
                snap.get("total_prompts"),
                snap.get("avg_prompts_per_user"),
                # Workloads
                snap.get("prompts_copilot_chat"),
                snap.get("prompts_teams"),
                snap.get("prompts_outlook"),
                snap.get("prompts_excel"),
                snap.get("prompts_word"),
                snap.get("prompts_powerpoint"),
                snap.get("prompts_onenote"),
                snap.get("prompts_loop"),
                # Agent Adoption
                snap.get("agent_adopters"),
                _pct(snap.get("agent_adoption_pct")),
                # Agent Inventory
                snap.get("total_agents"),
                snap.get("active_agents"),
                _pct(snap.get("utilization_rate")),
                snap.get("production_agents"),
                snap.get("non_prod_agents"),
                snap.get("agents_with_owner"),
                _pct(snap.get("ownership_pct")),
                snap.get("total_conversations"),
                # Environments
                snap.get("env_default"),
                snap.get("env_developer"),
                snap.get("env_teams"),
                snap.get("env_production"),
                snap.get("env_sandbox"),
                snap.get("env_trial"),
                # Entitlement
                round(total_entitled, 1)  if total_entitled  else None,
                round(total_prepaid, 1)   if total_prepaid   else None,
                round(total_payg, 1)      if total_payg      else None,
                round(remaining, 1)       if remaining       else None,
                pct_used,
                # Capacity Burn
                round(total_cap, 1) if total_cap else None,
                avg_daily,
                days_left,
            ]
            # Top LLM model credits
            for model in llm_headers:
                vals.append(round(model_totals.get(model, 0), 1) or None)

            fill = row_fill_even if row_idx % 2 == 0 else row_fill_odd
            for col, val in enumerate(vals, start=1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.fill      = fill
                cell.font      = data_font
                cell.alignment = left
                if col in _PCT_COLS:
                    cell.number_format = "0.0%"

    # ── Column widths ────────────────────────────────────────────────────────
    for col_idx, header in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        # Sample data rows to find max content width
        max_len = len(str(header))
        for row in ws.iter_rows(min_row=3, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 36)
