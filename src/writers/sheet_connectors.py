from openpyxl.worksheet.worksheet import Worksheet

from src.writers._style import apply_row_style, autofit_columns, write_headers

HEADERS = [
    "Timestamp",
    "gen_ai.operation.name",
    "gen_ai.tool.name",
    "Action Target",
    "gen_ai.agent.name",
    "gen_ai.agent.id",
    "gen_ai.environment.id",
    "Conversation ID",
    "User ID",
    "Channel",
    "Design Mode",
    "Success",
    "Result Code",
    "Duration (ms)",
]


def write(ws: Worksheet, connector_calls: list[dict]) -> None:
    calls = sorted(connector_calls, key=lambda c: c.get("Timestamp") or "", reverse=True)

    write_headers(ws, HEADERS)

    for row_idx, c in enumerate(calls, 2):
        ws.cell(row=row_idx, column=1, value=str(c.get("Timestamp", ""))[:19])
        ws.cell(row=row_idx, column=2, value=c.get("GenAiOperationName", ""))
        ws.cell(row=row_idx, column=3, value=c.get("ConnectorName", ""))
        ws.cell(row=row_idx, column=4, value=c.get("ActionTarget", ""))
        ws.cell(row=row_idx, column=5, value=c.get("GenAiAgentName", ""))
        ws.cell(row=row_idx, column=6, value=c.get("GenAiAgentId", ""))
        ws.cell(row=row_idx, column=7, value=c.get("GenAiEnvironmentId", ""))
        ws.cell(row=row_idx, column=8, value=c.get("ConversationId", ""))
        ws.cell(row=row_idx, column=9, value=c.get("UserId", ""))
        ws.cell(row=row_idx, column=10, value=c.get("ChannelId", ""))
        ws.cell(row=row_idx, column=11, value="Yes" if c.get("DesignMode") else "No")
        ws.cell(row=row_idx, column=12, value="Yes" if c.get("Success") else "No")
        ws.cell(row=row_idx, column=13, value=c.get("ResultCode", ""))
        ws.cell(row=row_idx, column=14, value=c.get("DurationMs", ""))
        apply_row_style(ws, row_idx, len(HEADERS))

    autofit_columns(ws, HEADERS)
