from openpyxl.worksheet.worksheet import Worksheet

from src.writers._style import autofit_columns, write_headers

HEADERS = [
    "Environment ID",
    "Display Name",
    "Type",
    "Region",
    "State",
    "Created Date",
    "Modified Date",
    "SKU",
    "Dataverse",
]


def write(ws: Worksheet, environments: list[dict]) -> None:
    write_headers(ws, HEADERS)

    if not environments:
        ws.cell(row=2, column=1, value="— No environment data synced yet —")
    else:
        for i, e in enumerate(environments, start=2):
            ws.cell(row=i, column=1, value=e.get("environment_id", ""))
            ws.cell(row=i, column=2, value=e.get("display_name", ""))
            ws.cell(row=i, column=3, value=e.get("type", ""))
            ws.cell(row=i, column=4, value=e.get("region", ""))
            ws.cell(row=i, column=5, value=e.get("state", ""))
            ws.cell(row=i, column=6, value=e.get("created_at", ""))
            ws.cell(row=i, column=7, value=e.get("modified_at", ""))
            ws.cell(row=i, column=8, value=e.get("sku", ""))
            ws.cell(row=i, column=9, value=e.get("dataverse_url", ""))

    autofit_columns(ws, HEADERS)
