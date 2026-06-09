from openpyxl.worksheet.worksheet import Worksheet

from src.writers._style import apply_row_style, autofit_columns, write_headers

HEADERS = [
    "Package ID", "Display Name", "Type", "State",
    "Publisher", "App ID", "Description",
]


def write(ws: Worksheet, rows: list[dict]) -> None:
    write_headers(ws, HEADERS)

    if not rows:
        ws.cell(row=2, column=1, value="— No Copilot packages found (requires Copilot.Read.All) —")
    else:
        for i, r in enumerate(rows, start=2):
            ws.cell(row=i, column=1, value=r.get("package_id", ""))
            ws.cell(row=i, column=2, value=r.get("display_name", ""))
            ws.cell(row=i, column=3, value=r.get("type", ""))
            ws.cell(row=i, column=4, value=r.get("state", ""))
            ws.cell(row=i, column=5, value=r.get("publisher_name", ""))
            ws.cell(row=i, column=6, value=r.get("app_id", ""))
            ws.cell(row=i, column=7, value=r.get("description", ""))
            apply_row_style(ws, i, len(HEADERS))

    autofit_columns(ws, HEADERS)
