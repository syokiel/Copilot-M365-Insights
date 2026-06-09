from collections import Counter
from datetime import datetime, timezone

from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from src.writers._style import HEADER_FILL, HEADER_FONT, LEFT, autofit_columns


def write(
    ws: Worksheet,
    events: list[dict],
    connector_calls: list[dict],
    model_calls: list[dict] = [],
    kpi_snapshot: dict | None = None,
    viva_cs_sessions: list[dict] | None = None,
    viva_cs_wau: list[dict] | None = None,
    viva_cs_autonomous: list[dict] | None = None,
    viva_cs_agents: dict[str, dict] | None = None,
) -> None:
    prod_events = [e for e in events if not e.get("DesignMode")]
    prod_connectors = [c for c in connector_calls if not c.get("DesignMode")]

    all_conversations = {e["ConversationId"] for e in events if e.get("ConversationId")}
    prod_conversations = {e["ConversationId"] for e in prod_events if e.get("ConversationId")}

    received = [e for e in events if e.get("EventName") == "BotMessageReceived"]
    sent = [e for e in events if e.get("EventName") == "BotMessageSend"]

    connector_counter: Counter = Counter(c.get("ConnectorName", "Unknown") for c in prod_connectors)
    failed_connectors = [c for c in prod_connectors if not c.get("Success")]

    timestamps = [e["Timestamp"] for e in events if e.get("Timestamp")]
    earliest = min(timestamps) if timestamps else None
    latest = max(timestamps) if timestamps else None

    total_input_tokens = sum(r.get("gen_ai_usage_input_tokens") or 0 for r in model_calls)
    total_output_tokens = sum(r.get("gen_ai_usage_output_tokens") or 0 for r in model_calls)
    model_counter: Counter = Counter(
        r.get("gen_ai_request_model") or r.get("gen_ai_response_model", "unknown")
        for r in model_calls
        if r.get("gen_ai_request_model") or r.get("gen_ai_response_model")
    )
    agents_using_ai = {r.get("gen_ai_agent_name") or r.get("gen_ai_agent_id", "") for r in model_calls if r.get("gen_ai_agent_id")}

    rows = [
        ("Report generated", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")),
        ("Data range (earliest)", str(earliest)[:19] if earliest else "—"),
        ("Data range (latest)", str(latest)[:19] if latest else "—"),
        (None, None),
        ("── Conversations ─────────────────────────────", None),
        ("Total conversations", len(all_conversations)),
        ("Production conversations", len(prod_conversations)),
        ("Design-mode conversations", len(all_conversations) - len(prod_conversations)),
        (None, None),
        ("── Messages ──────────────────────────────────", None),
        ("Total events", len(events)),
        ("Messages received (user → bot)", len(received)),
        ("Messages sent (bot → user)", len(sent)),
        (None, None),
        ("── Connector Calls (production) ──────────────", None),
        ("Total connector calls", len(prod_connectors)),
        ("Failed connector calls", len(failed_connectors)),
    ]

    if connector_counter:
        rows.append((None, None))
        rows.append(("── Top Connectors ────────────────────────────", None))
        for name, count in connector_counter.most_common(10):
            rows.append((f"  {name}", count))

    rows.append((None, None))
    rows.append(("── Generative AI Usage ───────────────────────", None))
    if model_calls:
        rows += [
            ("AI model calls",               len(model_calls)),
            ("Agents using generative AI",   len(agents_using_ai)),
            ("Total input tokens",           total_input_tokens),
            ("Total output tokens",          total_output_tokens),
            ("Total tokens",                 total_input_tokens + total_output_tokens),
        ]
        if model_counter:
            rows.append((None, None))
            rows.append(("── Models Used ───────────────────────────────", None))
            for model, count in model_counter.most_common():
                rows.append((f"  {model}", count))
    else:
        rows.append(("  No AI model call data in this window", None))

    if kpi_snapshot:
        pct = lambda v: f"{v:.1f}%" if v is not None else "—"  # noqa: E731
        rows += [
            (None, None),
            ("── M365 Copilot KPIs ─────────────────────────", None),
            ("Total Licenses",       kpi_snapshot.get("total_licenses") or "—"),
            ("Enabled Users",        kpi_snapshot.get("enabled_users")),
            ("Active Users",         kpi_snapshot.get("active_users")),
            ("Activation Rate",      pct(kpi_snapshot.get("activation_rate"))),
            ("Adoption Rate",        pct(kpi_snapshot.get("adoption_rate"))),
            ("Power Users",          kpi_snapshot.get("power_users")),
            ("Total Prompts",        kpi_snapshot.get("total_prompts")),
            ("Avg Prompts / User",   kpi_snapshot.get("avg_prompts_per_user")),
            (None, None),
            ("── Agent KPIs ────────────────────────────────", None),
            ("Total Agents",         kpi_snapshot.get("total_agents")),
            ("Active Agents",        kpi_snapshot.get("active_agents")),
            ("Utilization Rate",     pct(kpi_snapshot.get("utilization_rate"))),
            ("Production Agents",    kpi_snapshot.get("production_agents")),
            ("Non-Prod Agents",      kpi_snapshot.get("non_prod_agents")),
            ("Ownership Coverage",   pct(kpi_snapshot.get("ownership_pct"))),
            ("Total Conversations",  kpi_snapshot.get("total_conversations")),
            ("Agent Adopters",       kpi_snapshot.get("agent_adopters")),
            ("Agent Adoption %",     pct(kpi_snapshot.get("agent_adoption_pct"))),
        ]

    # ── Viva CS (Copilot Studio analytics) ────────────────────────────────────
    rows.append((None, None))
    rows.append(("── Viva CS (Copilot Studio) ──────────────────", None))

    if viva_cs_sessions:
        total_sess     = sum(r.get("total_sessions")    or 0 for r in viva_cs_sessions)
        total_engaged  = sum(r.get("engaged_sessions")  or 0 for r in viva_cs_sessions)
        total_resolved = sum(r.get("resolved_sessions") or 0 for r in viva_cs_sessions)
        total_escalated= sum(r.get("escalated_sessions")or 0 for r in viva_cs_sessions)
        total_abandoned= sum(r.get("abandoned_sessions")or 0 for r in viva_cs_sessions)
        total_csat_n   = sum(r.get("csat_responses")    or 0 for r in viva_cs_sessions)

        # Weighted avg CSAT
        csat_score = 0.0
        if total_csat_n:
            for r in viva_cs_sessions:
                n = r.get("csat_responses") or 0
                if n:
                    s = (
                        (r.get("csat_1") or 0) * 1 + (r.get("csat_2") or 0) * 2 +
                        (r.get("csat_3") or 0) * 3 + (r.get("csat_4") or 0) * 4 +
                        (r.get("csat_5") or 0) * 5
                    )
                    csat_score += s
            csat_score = round(csat_score / total_csat_n, 2)

        pct = lambda v, d: f"{v/d*100:.1f}%" if d else "—"  # noqa: E731

        rows += [
            ("Catalog agents",      len(viva_cs_agents) if viva_cs_agents else "—"),
            ("Total sessions",      total_sess),
            ("Engaged sessions",    f"{total_engaged}  ({pct(total_engaged, total_sess)})"),
            ("Resolved sessions",   f"{total_resolved}  ({pct(total_resolved, total_sess)})"),
            ("Escalated sessions",  f"{total_escalated}  ({pct(total_escalated, total_sess)})"),
            ("Abandoned sessions",  f"{total_abandoned}  ({pct(total_abandoned, total_sess)})"),
            ("CSAT responses",      total_csat_n),
            ("Avg CSAT score",      csat_score if total_csat_n else "—"),
        ]
    else:
        rows.append(("  No Copilot Studio session data imported", None))

    if viva_cs_wau:
        peak = max(viva_cs_wau, key=lambda r: r.get("active_user_count") or 0)
        rows += [
            ("Peak weekly active users", peak.get("active_user_count")),
            ("Peak WAU week",           peak.get("start_date", "—")),
        ]

    if viva_cs_autonomous:
        total_runs  = sum(r.get("total_runs")      or 0 for r in viva_cs_autonomous)
        total_succ  = sum(r.get("successful_runs") or 0 for r in viva_cs_autonomous)
        rows += [
            ("Autonomous runs (total)",   total_runs),
            ("Autonomous success rate",   f"{total_succ/total_runs*100:.1f}%" if total_runs else "—"),
        ]

    headers = ["Metric", "Value"]
    ws.column_dimensions["A"].width = 48
    ws.column_dimensions["B"].width = 28

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = LEFT

    ws.freeze_panes = "A2"

    for row_idx, (metric, value) in enumerate(rows, 2):
        a = ws.cell(row=row_idx, column=1, value=metric)
        b = ws.cell(row=row_idx, column=2, value=value)
        a.alignment = LEFT
        b.alignment = LEFT
        if metric and metric.startswith("──"):
            a.font = Font(bold=True, size=11, color="1F4E79")
        else:
            a.font = Font(size=11)
            b.font = Font(size=11)
