from collections import defaultdict
from datetime import datetime, timezone

from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from src.writers._style import HEADER_FILL, HEADER_FONT, LEFT

_WARN_FILL  = PatternFill("solid", fgColor="FFF2CC")   # amber — watch
_ALERT_FILL = PatternFill("solid", fgColor="FFE0E0")   # red   — action needed
_GOOD_FILL  = PatternFill("solid", fgColor="E2EFDA")   # green — healthy


def write(
    ws: Worksheet,
    entitlement: list[dict],
    per_agent: list[dict],
    per_user: list[dict],
    capacity: list[dict],
) -> None:

    # ── Aggregates: entitlement ───────────────────────────────────────────────
    total_entitled  = sum(r.get("entitled_quantity")          or 0 for r in entitlement)
    total_prepaid   = sum(r.get("prepaid_consumed_quantity")  or 0 for r in entitlement)
    total_payg      = sum(r.get("payg_consumed_quantity")     or 0 for r in entitlement)
    total_consumed  = total_prepaid + total_payg
    remaining       = max(0.0, total_entitled - total_prepaid)
    pct_used        = round(total_consumed / total_entitled * 100, 1) if total_entitled else 0.0

    usage_dates = [r.get("usage_date") for r in entitlement if r.get("usage_date")]
    date_min = min(usage_dates) if usage_dates else None
    date_max = max(usage_dates) if usage_dates else None

    # ── Aggregates: capacity burn rate ────────────────────────────────────────
    daily: dict[str, float] = defaultdict(float)
    billable_capacity   = 0.0
    unbillable_capacity = 0.0
    for r in capacity:
        qty = r.get("consumed_quantity") or 0
        d   = r.get("consumption_date", "")
        if d:
            daily[d[:10]] += qty
        if r.get("is_billable"):
            billable_capacity += qty
        else:
            unbillable_capacity += qty

    total_capacity = billable_capacity + unbillable_capacity
    avg_daily  = round(total_capacity / len(daily), 1) if daily else 0.0
    peak_day   = max(daily, key=daily.get) if daily else None
    peak_qty   = round(daily[peak_day], 1) if peak_day else 0.0
    days_left  = round(remaining / avg_daily) if avg_daily and remaining else None

    # ── Aggregates: per-agent credits ─────────────────────────────────────────
    agent_credits: dict[str, float] = defaultdict(float)
    agent_billed:  dict[str, float] = defaultdict(float)
    model_credits: dict[str, float] = defaultdict(float)
    channel_credits: dict[str, float] = defaultdict(float)
    feature_credits: dict[str, float] = defaultdict(float)

    for r in per_agent:
        name    = r.get("agent_name") or "Unknown"
        billed  = r.get("billed_credit")     or 0
        unbilled= r.get("non_billed_credit") or 0
        total   = billed + unbilled
        agent_credits[name]           += total
        agent_billed[name]            += billed
        model_credits[r.get("llm_model")   or "Unknown"] += total
        channel_credits[r.get("channel")   or "Unknown"] += total
        feature_credits[r.get("ai_feature")or "Unknown"] += total

    top_agents_total  = sorted(agent_credits.items(),  key=lambda x: x[1], reverse=True)[:10]
    top_agents_billed = sorted(agent_billed.items(),   key=lambda x: x[1], reverse=True)[:5]
    top_models        = sorted(model_credits.items(),  key=lambda x: x[1], reverse=True)
    top_channels      = sorted(channel_credits.items(),key=lambda x: x[1], reverse=True)
    top_features      = sorted(feature_credits.items(),key=lambda x: x[1], reverse=True)

    total_billed_credits   = sum(r.get("billed_credit")     or 0 for r in per_agent)
    total_unbilled_credits = sum(r.get("non_billed_credit") or 0 for r in per_agent)

    # ── Aggregates: per-user credits ──────────────────────────────────────────
    user_credits: dict[str, float] = defaultdict(float)
    user_billable: dict[str, float] = defaultdict(float)
    user_licensed: dict[str, bool]  = {}
    for r in per_user:
        email   = r.get("user_email") or r.get("user_id", "Unknown")
        credits = r.get("credits_used")        or 0
        billed  = r.get("billable_credit_used") or 0
        user_credits[email]  += credits
        user_billable[email] += billed
        if email not in user_licensed:
            user_licensed[email] = bool(r.get("m365_copilot_licensed"))

    top_users = sorted(user_credits.items(), key=lambda x: x[1], reverse=True)[:10]
    unlicensed_with_billable = [
        (email, round(user_billable[email], 1))
        for email, licensed in user_licensed.items()
        if not licensed and user_billable.get(email, 0) > 0
    ]
    unlicensed_with_billable.sort(key=lambda x: x[1], reverse=True)
    total_users_consuming = len(user_credits)
    licensed_credits   = sum(v for e, v in user_credits.items() if user_licensed.get(e))
    unlicensed_credits = sum(v for e, v in user_credits.items() if not user_licensed.get(e))

    # ── Environment-level entitlement ─────────────────────────────────────────
    env_totals: dict[str, dict] = {}
    for r in entitlement:
        env = r.get("environment_name") or "Unknown"
        if env not in env_totals:
            env_totals[env] = {"entitled": 0.0, "prepaid": 0.0, "payg": 0.0}
        env_totals[env]["entitled"] += r.get("entitled_quantity")         or 0
        env_totals[env]["prepaid"]  += r.get("prepaid_consumed_quantity") or 0
        env_totals[env]["payg"]     += r.get("payg_consumed_quantity")    or 0

    # ── Build row list ────────────────────────────────────────────────────────
    def _fmt(v) -> str:
        if v is None:
            return "—"
        if isinstance(v, float):
            return f"{v:,.1f}"
        return str(v)

    def _pct(v, d) -> str:
        return f"{v/d*100:.1f}%" if d else "—"

    rows: list[tuple] = [
        ("Report generated", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")),
        ("Data range", f"{date_min or '—'}  →  {date_max or '—'}"),
    ]

    # ── Section 1: Headline credit metrics ───────────────────────────────────
    rows += [
        (None, None),
        ("── Entitlement Overview ──────────────────────────", None),
        ("Total entitled credits",          _fmt(total_entitled)),
        ("Total prepaid consumed",          _fmt(total_prepaid)),
        ("Total PAYG consumed",             _fmt(total_payg)),
        ("Total consumed (prepaid + PAYG)", _fmt(total_consumed)),
        ("Remaining prepaid entitlement",   _fmt(remaining)),
        ("% of entitlement consumed",       f"{pct_used}%"),
        ("PAYG active?",                    "YES — review budget" if total_payg > 0 else "No"),
    ]

    # ── Section 2: Burn rate ──────────────────────────────────────────────────
    rows += [
        (None, None),
        ("── Capacity Burn Rate ────────────────────────────", None),
        ("Total capacity consumed",         _fmt(total_capacity)),
        ("  Billable",                      _fmt(billable_capacity)),
        ("  Non-billable",                  _fmt(unbillable_capacity)),
        ("Active consumption days",         str(len(daily))),
        ("Average daily consumption",       _fmt(avg_daily)),
        ("Peak single-day consumption",     f"{_fmt(peak_qty)}  ({peak_day or '—'})"),
        ("Estimated days of prepaid left",  str(days_left) if days_left is not None else "—"),
    ]

    # ── Section 3: Credit breakdown by agent ─────────────────────────────────
    rows += [
        (None, None),
        ("── Top Agents by Total Credits ──────────────────", None),
    ]
    for agent, total in top_agents_total:
        rows.append((f"  {agent}", _fmt(total)))

    rows += [
        (None, None),
        ("── Top Agents by Billed Credits ─────────────────", None),
    ]
    for agent, billed in top_agents_billed:
        rows.append((f"  {agent}", _fmt(billed)))

    # ── Section 4: LLM model breakdown ───────────────────────────────────────
    rows += [
        (None, None),
        ("── Credits by LLM Model ─────────────────────────", None),
    ]
    for model, total in top_models:
        rows.append((f"  {model}", _fmt(total)))

    # ── Section 5: Channel breakdown ──────────────────────────────────────────
    rows += [
        (None, None),
        ("── Credits by Channel ────────────────────────────", None),
    ]
    for channel, total in top_channels:
        rows.append((f"  {channel}", _fmt(total)))

    # ── Section 6: AI Feature breakdown ──────────────────────────────────────
    rows += [
        (None, None),
        ("── Credits by AI Feature ────────────────────────", None),
    ]
    for feature, total in top_features:
        rows.append((f"  {feature}", _fmt(total)))

    # ── Section 7: Per-user consumption ──────────────────────────────────────
    rows += [
        (None, None),
        ("── User Consumption ─────────────────────────────", None),
        ("Total users consuming credits",      str(total_users_consuming)),
        ("Credits from M365 licensed users",   _fmt(licensed_credits)),
        ("Credits from unlicensed users",      _fmt(unlicensed_credits)),
        ("Unlicensed users with billed credits", str(len(unlicensed_with_billable))),
        (None, None),
        ("── Top Users by Total Credits ───────────────────", None),
    ]
    for email, total in top_users:
        licensed_tag = "" if user_licensed.get(email) else "  ⚠ unlicensed"
        rows.append((f"  {email}{licensed_tag}", _fmt(total)))

    if unlicensed_with_billable:
        rows += [
            (None, None),
            ("── Unlicensed Users with Billed Credits ─────────", None),
        ]
        for email, billed in unlicensed_with_billable[:10]:
            rows.append((f"  {email}", _fmt(billed)))

    # ── Section 8: Environment-level entitlement status ───────────────────────
    rows += [
        (None, None),
        ("── Entitlement by Environment ───────────────────", None),
    ]
    for env, t in sorted(env_totals.items()):
        pct = round((t["prepaid"] + t["payg"]) / t["entitled"] * 100, 1) if t["entitled"] else 0.0
        payg_flag = "  ⚠ PAYG active" if t["payg"] > 0 else ""
        rows.append((f"  {env}", f"{pct}% used  |  prepaid {_fmt(t['prepaid'])}  /  {_fmt(t['entitled'])}{payg_flag}"))

    # ── Write headers ─────────────────────────────────────────────────────────
    for col, header in enumerate(["Metric", "Value"], 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = LEFT

    ws.freeze_panes  = "A2"
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 40

    # ── Write rows ────────────────────────────────────────────────────────────
    for row_idx, (metric, value) in enumerate(rows, 2):
        a = ws.cell(row=row_idx, column=1, value=metric)
        b = ws.cell(row=row_idx, column=2, value=value)
        a.alignment = LEFT
        b.alignment = LEFT

        if metric and metric.startswith("──"):
            a.font = Font(bold=True, size=11, color="1F4E79")
            b.font = Font(size=11)
        else:
            a.font = Font(size=11)
            b.font = Font(size=11)

        # Colour-code alert rows
        val_str = str(value or "")
        if metric == "PAYG active?" and "YES" in val_str:
            a.fill = _ALERT_FILL
            b.fill = _ALERT_FILL
        elif metric == "Unlicensed users with billed credits" and value not in (None, "0", "—"):
            a.fill = _WARN_FILL
            b.fill = _WARN_FILL
        elif metric == "Estimated days of prepaid left" and value not in (None, "—"):
            try:
                if int(value) < 30:
                    a.fill = _ALERT_FILL
                    b.fill = _ALERT_FILL
                elif int(value) < 90:
                    a.fill = _WARN_FILL
                    b.fill = _WARN_FILL
                else:
                    a.fill = _GOOD_FILL
                    b.fill = _GOOD_FILL
            except (ValueError, TypeError):
                pass
        elif metric == "% of entitlement consumed":
            try:
                pct_v = float(val_str.replace("%", ""))
                if pct_v >= 90:
                    a.fill = _ALERT_FILL
                    b.fill = _ALERT_FILL
                elif pct_v >= 75:
                    a.fill = _WARN_FILL
                    b.fill = _WARN_FILL
                else:
                    a.fill = _GOOD_FILL
                    b.fill = _GOOD_FILL
            except (ValueError, TypeError):
                pass
