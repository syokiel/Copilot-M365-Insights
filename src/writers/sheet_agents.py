from openpyxl.worksheet.worksheet import Worksheet

from src.writers._style import autofit_columns, write_headers

HEADERS = [
    "Agent ID",
    "Agent Name",
    "Environment ID",
    "Environment Name",
    "Created By",
    "Created Date",
    "Modified Date",
    "Status",
    "Channel",
    "Solution",
]


def write(ws: Worksheet, agents: list[dict]) -> None:
    write_headers(ws, HEADERS)

    if not agents:
        ws.cell(row=2, column=1, value="— No agent data synced yet —")
    else:
        for row_idx, bot in enumerate(agents, start=2):
            published = bot.get("published_at") or bot.get("publishedDateTime", "")
            ws.cell(row=row_idx, column=1, value=bot.get("bot_id") or bot.get("id", ""))
            ws.cell(row=row_idx, column=2, value=bot.get("display_name") or bot.get("name", ""))
            ws.cell(row=row_idx, column=3, value=bot.get("environment_id") or bot.get("environmentId", ""))
            ws.cell(row=row_idx, column=4, value="")  # Environment Name — not available
            ws.cell(row=row_idx, column=5, value="")  # Created By — not available
            ws.cell(row=row_idx, column=6, value=bot.get("created_at") or bot.get("createdDateTime", ""))
            ws.cell(row=row_idx, column=7, value=bot.get("modified_at") or bot.get("modifiedDateTime", ""))
            ws.cell(row=row_idx, column=8, value="Published" if published else "Draft")
            ws.cell(row=row_idx, column=9, value="")  # Channel — not available
            ws.cell(row=row_idx, column=10, value=bot.get("schema_name") or bot.get("schemaName", ""))

    autofit_columns(ws, HEADERS)
