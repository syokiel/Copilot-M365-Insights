from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
ROW_FILL_ODD = PatternFill("solid", fgColor="FFFFFF")
ROW_FILL_EVEN = PatternFill("solid", fgColor="D6E4F0")
DATA_FONT = Font(size=11)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=False)


def write_headers(ws: Worksheet, headers: list[str]) -> None:
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 20


def apply_row_style(ws: Worksheet, row: int, col_count: int) -> None:
    fill = ROW_FILL_EVEN if row % 2 == 0 else ROW_FILL_ODD
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = DATA_FONT
        cell.alignment = LEFT


def autofit_columns(ws: Worksheet, headers: list[str], max_width: int = 50) -> None:
    for col, header in enumerate(headers, 1):
        col_letter = ws.cell(row=1, column=col).column_letter
        max_len = len(header)
        for row in ws.iter_rows(min_row=2, min_col=col, max_col=col):
            for cell in row:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)
