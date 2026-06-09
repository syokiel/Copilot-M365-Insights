from openpyxl.worksheet.worksheet import Worksheet

from src.writers._style import apply_row_style, autofit_columns, write_headers

HEADERS = [
    "Agent ID", "Agent Name", "Metric Date",
    "Total Sessions", "Engaged", "Unengaged",
    "Resolved", "Escalated", "Abandoned",
    "Engagement %", "Resolution %", "Escalation %", "Abandonment %",
    "CSAT Responses", "Avg CSAT",
    "Avg Duration (All s)", "Avg Duration (Engaged s)",
    "KS Sessions (Engaged)", "KS Sessions (Resolved)",
]


def _pct(num, denom) -> str | None:
    if denom and denom > 0 and num is not None:
        return round(num / denom * 100, 1)
    return None


def _avg_csat(r: dict) -> float | None:
    total = r.get('csat_responses') or 0
    if total == 0:
        return None
    score = (
        (r.get('csat_1') or 0) * 1 +
        (r.get('csat_2') or 0) * 2 +
        (r.get('csat_3') or 0) * 3 +
        (r.get('csat_4') or 0) * 4 +
        (r.get('csat_5') or 0) * 5
    )
    return round(score / total, 2)


def write(ws: Worksheet, rows: list[dict], agents: dict[str, dict] | None = None) -> None:
    write_headers(ws, HEADERS)
    ag = agents or {}

    if not rows:
        ws.cell(row=2, column=1, value="— No session metrics imported —")
    else:
        for i, r in enumerate(rows, start=2):
            aid = r.get('agent_id', '')
            total = r.get('total_sessions')
            ws.cell(row=i, column=1,  value=aid)
            ws.cell(row=i, column=2,  value=ag.get(aid, {}).get('agent_name', ''))
            ws.cell(row=i, column=3,  value=r.get('metric_date', ''))
            ws.cell(row=i, column=4,  value=total)
            ws.cell(row=i, column=5,  value=r.get('engaged_sessions'))
            ws.cell(row=i, column=6,  value=r.get('unengaged_sessions'))
            ws.cell(row=i, column=7,  value=r.get('resolved_sessions'))
            ws.cell(row=i, column=8,  value=r.get('escalated_sessions'))
            ws.cell(row=i, column=9,  value=r.get('abandoned_sessions'))
            ws.cell(row=i, column=10, value=_pct(r.get('engaged_sessions'), total))
            ws.cell(row=i, column=11, value=_pct(r.get('resolved_sessions'), total))
            ws.cell(row=i, column=12, value=_pct(r.get('escalated_sessions'), total))
            ws.cell(row=i, column=13, value=_pct(r.get('abandoned_sessions'), total))
            ws.cell(row=i, column=14, value=r.get('csat_responses'))
            ws.cell(row=i, column=15, value=_avg_csat(r))
            ws.cell(row=i, column=16, value=r.get('avg_duration_all'))
            ws.cell(row=i, column=17, value=r.get('avg_duration_engaged'))
            ws.cell(row=i, column=18, value=r.get('ks_engaged'))
            ws.cell(row=i, column=19, value=r.get('ks_resolved'))
            apply_row_style(ws, i, len(HEADERS))

    autofit_columns(ws, HEADERS)
