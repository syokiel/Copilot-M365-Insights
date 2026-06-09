from openpyxl.worksheet.worksheet import Worksheet

from src.writers._style import apply_row_style, autofit_columns, write_headers

HEADERS = [
    "User Principal Name", "Display Name",
    "Exchange Last Activity", "Teams Last Activity",
    "SharePoint Last Activity", "OneDrive Last Activity",
    "Has Exchange", "Has Teams", "Has SharePoint", "Has OneDrive",
    "Report Refresh Date",
]


def write(ws: Worksheet, rows: list[dict]) -> None:
    write_headers(ws, HEADERS)

    if not rows:
        ws.cell(row=2, column=1, value="— Requires Reports.Read.All permission —")
    else:
        for i, r in enumerate(rows, start=2):
            ws.cell(row=i, column=1,  value=r.get("user_principal_name", ""))
            ws.cell(row=i, column=2,  value=r.get("display_name", ""))
            ws.cell(row=i, column=3,  value=r.get("exchange_last_activity", ""))
            ws.cell(row=i, column=4,  value=r.get("teams_last_activity", ""))
            ws.cell(row=i, column=5,  value=r.get("sharepoint_last_activity", ""))
            ws.cell(row=i, column=6,  value=r.get("onedrive_last_activity", ""))
            ws.cell(row=i, column=7,  value="Yes" if r.get("has_exchange_license") else "No")
            ws.cell(row=i, column=8,  value="Yes" if r.get("has_teams_license") else "No")
            ws.cell(row=i, column=9,  value="Yes" if r.get("has_sharepoint_license") else "No")
            ws.cell(row=i, column=10, value="Yes" if r.get("has_onedrive_license") else "No")
            ws.cell(row=i, column=11, value=r.get("report_refresh_date", ""))
            apply_row_style(ws, i, len(HEADERS))

    autofit_columns(ws, HEADERS)
