from openpyxl.worksheet.worksheet import Worksheet

from src.writers._style import apply_row_style, autofit_columns, write_headers

HEADERS = [
    "User ID",
    "Display Name",
    "UPN",
    "Week Start",
    "Week End",
    "Focus Hours",
    "Meeting Hours",
    "Email Hours",
    "Chat Hours",
    "After Hours",
]


def write(ws: Worksheet, rows: list[dict], aad_users: dict[str, dict] | None = None) -> None:
    write_headers(ws, HEADERS)
    aad = aad_users or {}

    if not rows:
        ws.cell(row=2, column=1, value="— Requires Analytics.ReadAll permission on the sync service principal —")
    else:
        for i, r in enumerate(rows, start=2):
            uid = r.get("user_id", "")
            user = aad.get(uid, {})
            ws.cell(row=i, column=1, value=uid)
            ws.cell(row=i, column=2, value=user.get("display_name", ""))
            ws.cell(row=i, column=3, value=user.get("upn", ""))
            ws.cell(row=i, column=4, value=r.get("week_start", ""))
            ws.cell(row=i, column=5, value=r.get("week_end", ""))
            ws.cell(row=i, column=6, value=r.get("focus_hours"))
            ws.cell(row=i, column=7, value=r.get("meeting_hours"))
            ws.cell(row=i, column=8, value=r.get("email_hours"))
            ws.cell(row=i, column=9, value=r.get("chat_hours"))
            ws.cell(row=i, column=10, value=r.get("after_hours"))
            apply_row_style(ws, i, len(HEADERS))

    autofit_columns(ws, HEADERS)
