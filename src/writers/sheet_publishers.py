from openpyxl.worksheet.worksheet import Worksheet

from src.writers._style import autofit_columns, write_headers

HEADERS = [
    "Publisher ID",
    "Display Name",
    "Unique Name",
    "Email",
    "Phone",
    "Custom Prefix",
    "Solution Count",
]


def write(ws: Worksheet, publishers: list[dict]) -> None:
    write_headers(ws, HEADERS)

    if not publishers:
        ws.cell(row=2, column=1, value="— No publisher data synced yet —")
    else:
        for i, p in enumerate(publishers, start=2):
            ws.cell(row=i, column=1, value=p.get("publisher_id", ""))
            ws.cell(row=i, column=2, value=p.get("display_name", ""))
            ws.cell(row=i, column=3, value=p.get("unique_name", ""))
            ws.cell(row=i, column=4, value=p.get("email", ""))
            ws.cell(row=i, column=5, value=p.get("phone", ""))
            ws.cell(row=i, column=6, value=p.get("custom_prefix", ""))
            ws.cell(row=i, column=7, value=p.get("solution_count"))

    autofit_columns(ws, HEADERS)
