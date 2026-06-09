from openpyxl.worksheet.worksheet import Worksheet

from src.writers._style import apply_row_style, autofit_columns, write_headers

HEADERS = [
    "User Principal Name",
    "Last Activation Date", "Last Activity Date",
    "Outlook", "Word", "Excel", "PowerPoint",
    "OneNote", "Teams", "SharePoint", "OneDrive",
    "Report Refresh Date",
]

_APP_FIELDS = [
    "outlook_active", "word_active", "excel_active", "ppt_active",
    "onenote_active", "teams_active", "sharepoint_active", "onedrive_active",
]


def _yn(v) -> str:
    if v is None:
        return ""
    return "Yes" if v else "No"


def write(ws: Worksheet, rows: list[dict]) -> None:
    write_headers(ws, HEADERS)

    if not rows:
        ws.cell(row=2, column=1, value="— Requires Reports.Read.All permission —")
    else:
        for i, r in enumerate(rows, start=2):
            ws.cell(row=i, column=1,  value=r.get("user_principal_name", ""))
            ws.cell(row=i, column=2,  value=r.get("last_activation_date", ""))
            ws.cell(row=i, column=3,  value=r.get("last_activity_date", ""))
            for col_offset, field in enumerate(_APP_FIELDS):
                ws.cell(row=i, column=4 + col_offset, value=_yn(r.get(field)))
            ws.cell(row=i, column=12, value=r.get("report_refresh_date", ""))
            apply_row_style(ws, i, len(HEADERS))

    autofit_columns(ws, HEADERS)
